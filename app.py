from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


APP_TITLE = "BIW DFMEA-DVP&R AI Assistant"
TOOL_VERSION = "MVP-0.6"
PARTS_DIR = Path(__file__).parent / "examples" / "parts"
DISCLAIMER = (
    "Draft engineering content for review only. All DFMEA, DVP&R, rating, "
    "validation, and release decisions require responsible engineering approval."
)
PROTOTYPE_BOUNDARY_STATEMENT = (
    "This prototype uses synthetic BIW example data and rule-based AI logic for demonstration. "
    "It does not use confidential program data. A production version would need to run in a secure "
    "internal environment connected to approved DFMEA, DVP&R, CAE, validation, lessons learned, "
    "and engineering standards documents."
)
ACTION_PRIORITY_DISCLAIMER = (
    "Prototype Action Priority logic is used for demonstration. Production deployment should align "
    "Action Priority calculation with the company's approved FMEA rating method, internal quality "
    "standards, and customer-specific requirements."
)
GAP_WORKFLOW_NOTE = (
    "A validation gap remains open until the responsible engineer accepts the AI-recommended closure "
    "action and links supporting validation evidence."
)
METHODOLOGY_NOTE = (
    "Structure follows the AIAG-VDA 7-step FMEA approach: planning, structure analysis, function "
    "analysis (P-Diagram), failure analysis, risk analysis (S/O/D with AIAG-VDA Action Priority and "
    "legacy RPN), optimization (recommended actions and revised ratings), and results documentation "
    "(traceability, gap analysis, and export)."
)
SPECIAL_CHARACTERISTIC_VALUES = ["", "YC (Potential Critical)", "SC (Potential Significant)"]
TEST_STAGE_VALUES = ["Virtual / CAE", "DV", "PV", "DV / PV"]
FINAL_PITCH_POSITIONING = (
    "This is not a production DFMEA/DVP&R replacement. This is a working proof of concept. "
    "The goal is to test whether AI can help BIW engineers generate first-draft DFMEA content, "
    "recommend DVP&R validation items, calculate risk priority, link risks to validation coverage, "
    "and flag gaps for engineer review. The engineer remains the final decision-maker."
)
DEMO_STORY = (
    "For the front rail reinforcement example, the tool generated 9 DFMEA failure modes, identified "
    "4 high-severity risks, recommended or proposed 14 DVP&R validation items, calculated initial "
    "and revised RPN, estimated 43% residual risk reduction, reused 8 lessons learned, and identified "
    "1 high-priority validation gap with a proposed closure test."
)
DASHBOARD_FOOTNOTES = [
    "RPN = Severity x Occurrence x Detection.",
    "Revised RPN estimates residual risk after proposed mitigation.",
    "Coverage Score reflects linked validation maturity.",
    "AI recommendations require engineer review before release.",
    "Synthetic demo data used for prototype only.",
]
PITCH_READINESS_ITEMS = [
    "Dashboard KPI labels are clear",
    "Percentages are formatted correctly",
    "Chart legends are renamed",
    "DVP&R test IDs are sorted",
    "Open gaps are clearly marked",
    "AI-proposed closure tests are visible",
    "Prototype disclaimer is visible",
    "Engineer approval boundary is visible",
    "Action Priority disclaimer is included",
    "Management demo story is included",
    "File does not contain confidential company data",
]

COMPONENT_FIELDS = [
    "component_name",
    "vehicle_area",
    "component_category",
    "material",
    "thickness",
    "joining_method",
    "manufacturing_process",
    "primary_function",
    "interfaces",
    "load_cases",
    "environmental_exposure",
    "known_design_concerns",
    "assumptions",
    "program_phase",
    "engineer_name",
]

GENERATION_SETTINGS = ["demo_gap_mode"]

DFMEA_COLUMNS = [
    "Failure Mode ID",
    "Function ID",
    "Requirement ID",
    "Action ID",
    "Item",
    "Function",
    "Risk Category",
    "Potential Failure Mode",
    "Potential Effect of Failure",
    "End Effect (Vehicle Level)",
    "Special Characteristic",
    "Potential Cause / Mechanism",
    "Initial Severity",
    "Initial Occurrence",
    "Initial Detection",
    "Initial RPN",
    "Severity",
    "Occurrence",
    "Detection",
    "RPN",
    "Action Priority",
    "AP (AIAG-VDA)",
    "Prevention Control",
    "Detection Control",
    "Recommended Action",
    "Revised Severity",
    "Revised Occurrence",
    "Revised Detection",
    "Revised RPN",
    "RPN Reduction %",
    "Residual Risk Level",
    "Action Owner",
    "Responsible Team",
    "Target Completion Date",
    "Action Status",
    "Closure Evidence",
    "Closure Notes",
    "Notes / Rationale",
    "AI Suggestion ID",
    "AI Confidence Score",
    "Source Type",
    "Source Strength",
    "Citation Count",
    "Human Review Required",
    "Reason for Human Review",
    "Future RAG Citation",
    "Source Evidence",
    "Source File",
    "Source Sheet",
    "Source Row",
    "Source Chunk ID",
    "AI Confidence",
    "Review Status",
    "Engineer Decision",
    "Rejection Reason",
    "Final Approved Text",
    "Reviewed By",
    "Review Date",
    "Approval Status",
    "Approval Notes",
    "Change Log",
]

DVP_COLUMNS = [
    "Test ID",
    "Requirement ID",
    "Linked Failure Mode ID",
    "Linked Failure Mode",
    "Recommended Validation Test",
    "Validation Type",
    "Validation Level",
    "Test Objective",
    "Test Method / Procedure",
    "Test Standard Reference",
    "Test Stage",
    "Build Phase",
    "Sample Size",
    "Test Duration / Cycles",
    "Test Conditions",
    "Acceptance Criteria",
    "Planned Start Date",
    "Planned Completion Date",
    "Actual Completion Date",
    "Actual Result",
    "Pass / Fail",
    "Issue ID",
    "Evidence Link",
    "Validation Status",
    "Responsible Team",
    "Test Owner",
    "Notes",
    "AI Suggestion ID",
    "AI Confidence Score",
    "Source Type",
    "Source Strength",
    "Citation Count",
    "Human Review Required",
    "Reason for Human Review",
    "Future RAG Citation",
    "Source Evidence",
    "Source File",
    "Source Sheet",
    "Source Row",
    "Source Chunk ID",
    "AI Confidence",
    "Review Status",
    "Engineer Decision",
    "Rejection Reason",
    "Final Approved Text",
    "Reviewed By",
    "Review Date",
    "Approval Status",
    "Approval Notes",
    "Change Log",
]

TRACE_COLUMNS = [
    "Requirement ID",
    "Function ID",
    "Failure Mode ID",
    "Failure Mode",
    "Risk Category",
    "Severity",
    "RPN",
    "Action Priority",
    "Linked Test IDs",
    "Linked DVP&R Test",
    "Coverage Status",
    "Coverage Score",
    "Coverage Reason",
    "Missing Validation Type",
    "Recommended Additional Test",
    "AI Recommended Test ID",
    "AI Recommended Test Name",
    "Before AI Review Coverage Status",
    "After AI Recommendation Coverage Status",
    "Expected Coverage Improvement",
    "Gap Description",
    "Recommended Action",
    "Priority",
    "Owner",
    "AI Suggestion ID",
    "AI Confidence Score",
    "Source Type",
    "Source Strength",
    "Citation Count",
    "Human Review Required",
    "Reason for Human Review",
    "Future RAG Citation",
    "Engineer Decision",
    "Rejection Reason",
    "Final Approved Text",
    "Reviewed By",
    "Review Date",
    "Approval Status",
    "Approval Notes",
    "Change Log",
]

GAP_COLUMNS = [
    "Gap ID",
    "Failure Mode ID",
    "Failure Mode",
    "Severity",
    "RPN",
    "Action Priority",
    "Risk Category",
    "Gap Type",
    "Gap Description",
    "Recommended Fix",
    "Responsible Team",
    "Priority",
    "Status",
    "Gap Status",
    "AI Recommended Test ID",
    "AI Recommended Test Name",
    "AI Recommended Closure Test",
    "Expected Coverage Improvement",
    "After Engineer Approval Coverage Status",
    "Engineer Decision",
    "Gap Closure Status",
    "Evidence Link",
    "Final Gap Closure Status",
    "Rejection Reason",
    "Final Approved Text",
    "Reviewed By",
    "Review Date",
    "Approval Status",
    "Change Log",
]

LESSON_COLUMNS = [
    "Lesson ID",
    "Related Failure Mode ID",
    "Related Requirement ID",
    "Risk Category",
    "Related Failure Mode",
    "Lesson Learned",
    "Prior Issue Type",
    "Applicability to Current Component",
    "Recommended Design Action",
    "Recommended Validation Action",
    "Source Type",
    "Source Confidence",
    "Engineer Decision",
    "AI Suggestion ID",
    "AI Confidence Score",
    "Source Strength",
    "Citation Count",
    "Human Review Required",
    "Reason for Human Review",
    "Future RAG Citation",
    "Source Evidence",
    "Source File",
    "Source Sheet",
    "Source Row",
    "Source Chunk ID",
    "AI Confidence",
    "Review Status",
    "Rejection Reason",
    "Final Approved Text",
    "Reviewed By",
    "Review Date",
    "Approval Status",
    "Approval Notes",
    "Change Log",
]

P_DIAGRAM_COLUMNS = ["Element", "Category", "Description"]

ACTION_STATUS_VALUES = ["Open", "In Progress", "Complete", "Rejected", "Deferred"]
VALIDATION_LEVEL_VALUES = ["CAE", "Coupon", "Component", "Subsystem", "Vehicle", "Subsystem / Vehicle"]
BUILD_PHASE_VALUES = ["Concept", "Mule", "Alpha", "Beta", "Production Intent", "Launch"]
PASS_FAIL_VALUES = ["TBD", "Not Run", "Pass", "Fail", "Waived"]
VALIDATION_STATUS_VALUES = ["Planned", "Proposed", "In Progress", "Passed", "Failed", "Waived", "Blocked"]
GAP_STATUS_VALUES = ["Open", "Proposed", "Accepted", "In Progress", "Closed", "Waived"]
ENGINEER_DECISION_VALUES = ["Accept", "Modify", "Reject", "Pending", "Needs More Data"]
APPROVAL_STATUS_VALUES = ["Draft", "Under Review", "Approved", "Rejected", "Deferred"]
REJECTION_REASON_VALUES = ["Not applicable", "Duplicate", "Wrong risk", "Needs more data", "Other"]
RISK_CATEGORIES = [
    "Crash / Safety",
    "Durability",
    "Corrosion",
    "Joining",
    "Joining / Durability",
    "Stamping / Formability",
    "Dimensional",
    "NVH",
    "Assembly",
    "Serviceability",
    "Manufacturing Quality",
    "Attachment / Durability",
]


@dataclass(frozen=True)
class RuleContext:
    inputs: dict[str, str]
    text: str

    def has_any(self, terms: list[str]) -> bool:
        return any(term.lower() in self.text for term in terms)


def empty_df(columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=columns)


def load_part_examples() -> dict[str, dict[str, str]]:
    examples: dict[str, dict[str, str]] = {}
    if not PARTS_DIR.exists():
        return examples

    for path in sorted(PARTS_DIR.glob("*.json")):
        with path.open("r", encoding="utf-8") as file:
            raw = json.load(file)
        inputs = {field: str(raw.get(field, "")).strip() for field in COMPONENT_FIELDS}
        label = inputs.get("component_name") or path.stem.replace("_", " ").title()
        examples[label] = inputs

    return examples


PART_EXAMPLES = load_part_examples()
DEFAULT_PART_LABEL = next(iter(PART_EXAMPLES), "")
DEMO_INPUTS = PART_EXAMPLES.get(DEFAULT_PART_LABEL, {})


# --------------------------------------------------------------------------- RAG
try:
    from rag import VectorStore, load_file_to_chunks
    from rag import config as rag_config
    from rag.llm import generate_llm_enrichment, llm_status
    from rag.prompt_builder import build_rag_prompt
    from rag.retriever import best_match_for_row, citation_label, retrieve_groups

    RAG_IMPORT_ERROR = ""
    VECTOR_DB_PATH = Path(rag_config.VECTOR_DB_PATH)
    if not VECTOR_DB_PATH.is_absolute():
        VECTOR_DB_PATH = Path(__file__).parent / rag_config.VECTOR_DB_PATH
except Exception as _rag_exc:  # keeps the app usable if RAG deps are missing
    RAG_IMPORT_ERROR = str(_rag_exc)
    VECTOR_DB_PATH = Path(__file__).parent / "data" / "vector_store"

KB_SEED_DIR = Path(__file__).parent / "data" / "knowledge_base"
KB_UPLOAD_DIR = Path(__file__).parent / "data" / "uploaded_docs"
RAG_DISCLAIMER = (
    "RAG results are used as engineering reference only. Engineer review and approval "
    "are required before release."
)
RAG_FALLBACK_LABEL = "No RAG source found - rule-based draft"
RAG_DOC_TYPES = [
    "Historical DFMEA",
    "Historical DVP&R",
    "Lessons Learned",
    "BIW Design Standard",
    "Weld Standard",
    "Corrosion Standard",
    "CAE Report",
    "Validation Report",
    "Launch Issue Report",
    "Manufacturing Feasibility Report",
    "Other",
]
RAG_SOURCE_STRENGTHS = [
    "Synthetic Demo",
    "Historical Program Document",
    "Engineering Standard",
    "Validation Report",
    "CAE Report",
    "Lessons Learned",
    "Unknown",
]
RETRIEVED_SOURCES_COLUMNS = [
    "Query ID",
    "Generation Type",
    "Source Rank",
    "Similarity Score",
    "Document Type",
    "Source File",
    "Source Sheet",
    "Source Row",
    "Source Chunk ID",
    "Chunk Preview",
    "Used In Output",
]
_SEED_FILE_TYPES = {
    "historical_dfmea_biw.xlsx": "Historical DFMEA",
    "historical_dvpr_biw.xlsx": "Historical DVP&R",
    "lessons_learned_biw.xlsx": "Lessons Learned",
    "biw_weld_standard_ws_join_003.md": "Weld Standard",
    "biw_corrosion_standard_ts_cor_021.md": "Corrosion Standard",
    "biw_dimensional_standard_dim_build_007.md": "BIW Design Standard",
    "launch_issue_reports_biw.md": "Launch Issue Report",
}


@st.cache_resource
def get_rag_store() -> "VectorStore":
    return VectorStore(VECTOR_DB_PATH)


def seed_knowledge_base(store: "VectorStore") -> int:
    """Index the bundled synthetic corpus. Duplicate chunks are skipped by hash."""
    added = 0
    if not KB_SEED_DIR.exists():
        return added
    for file in sorted(KB_SEED_DIR.iterdir()):
        doc_type = _SEED_FILE_TYPES.get(file.name, "Other")
        chunks = load_file_to_chunks(file, doc_type, "Synthetic Demo")
        added += store.add_chunks(chunks)
    return added


def _confidence_from_similarity(similarity: float) -> float:
    return round(min(0.95, 0.5 + similarity / 2), 2)


def _ground_frame(
    df: pd.DataFrame,
    text_columns: list[str],
    document_type: str,
    generation_type: str,
    store: "VectorStore",
    retrieved_log: list[dict[str, Any]],
) -> pd.DataFrame:
    """Attach best-match citations from the knowledge base to generated rows."""
    if df.empty:
        return df
    df = df.copy()
    for idx, row in df.iterrows():
        row_text = " ".join(str(row.get(col, "")) for col in text_columns)
        match = best_match_for_row(store, row_text, document_type)
        if match is None:
            if "Future RAG Citation" in df.columns:
                df.at[idx, "Future RAG Citation"] = RAG_FALLBACK_LABEL
            if "Source Evidence" in df.columns:
                df.at[idx, "Source Evidence"] = RAG_FALLBACK_LABEL
            continue
        meta = match.get("metadata", {})
        confidence = _confidence_from_similarity(match["similarity"])
        if "Future RAG Citation" in df.columns:
            df.at[idx, "Future RAG Citation"] = citation_label(match)
        if "Source Type" in df.columns:
            df.at[idx, "Source Type"] = "RAG Retrieved (rules draft)"
        if "Source Strength" in df.columns:
            df.at[idx, "Source Strength"] = meta.get("source_strength", "Unknown")
        if "Citation Count" in df.columns:
            df.at[idx, "Citation Count"] = 1
        if "AI Confidence Score" in df.columns:
            df.at[idx, "AI Confidence Score"] = confidence
        # Roadmap source schema
        if "Source Evidence" in df.columns:
            df.at[idx, "Source Evidence"] = str(match.get("chunk_text", ""))[:180]
        if "Source File" in df.columns:
            df.at[idx, "Source File"] = meta.get("file_name", "")
        if "Source Sheet" in df.columns:
            df.at[idx, "Source Sheet"] = meta.get("sheet_name", "")
        if "Source Row" in df.columns:
            df.at[idx, "Source Row"] = meta.get("row_number", "")
        if "Source Chunk ID" in df.columns:
            df.at[idx, "Source Chunk ID"] = match.get("chunk_id", "")
        if "AI Confidence" in df.columns:
            df.at[idx, "AI Confidence"] = confidence
        if "Review Status" in df.columns:
            df.at[idx, "Review Status"] = "Draft - source grounded"
        retrieved_log.append(
            {
                "Query ID": f"Q-{generation_type[:5].upper()}-{len(retrieved_log) + 1:03d}",
                "Generation Type": generation_type,
                "Source Rank": 1,
                "Similarity Score": match["similarity"],
                "Document Type": meta.get("document_type", ""),
                "Source File": meta.get("file_name", ""),
                "Source Sheet": meta.get("sheet_name", ""),
                "Source Row": meta.get("row_number", ""),
                "Source Chunk ID": match.get("chunk_id", ""),
                "Chunk Preview": str(match.get("chunk_text", ""))[:180],
                "Used In Output": "Yes",
            }
        )
    return df


def ground_with_rag(
    dfmea_df: pd.DataFrame,
    dvp_df: pd.DataFrame,
    lessons_df: pd.DataFrame,
    inputs: dict[str, str],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Ground generated drafts with knowledge-base citations. Never breaks generation."""
    empty_log = empty_df(RETRIEVED_SOURCES_COLUMNS)
    if RAG_IMPORT_ERROR:
        return dfmea_df, dvp_df, lessons_df, empty_log
    try:
        store = get_rag_store()
        if store.get_collection_stats()["chunks"] == 0:
            return dfmea_df, dvp_df, lessons_df, empty_log
        retrieved_log: list[dict[str, Any]] = []
        dfmea_df = _ground_frame(
            dfmea_df,
            ["Potential Failure Mode", "Potential Effect of Failure", "Potential Cause / Mechanism", "Risk Category"],
            "Historical DFMEA",
            "DFMEA",
            store,
            retrieved_log,
        )
        dvp_df = _ground_frame(
            dvp_df,
            ["Recommended Validation Test", "Linked Failure Mode", "Test Objective", "Acceptance Criteria"],
            "Historical DVP&R",
            "DVP&R",
            store,
            retrieved_log,
        )
        lessons_df = _ground_frame(
            lessons_df,
            ["Related Failure Mode", "Lesson Learned", "Risk Category"],
            "Lessons Learned",
            "Lessons",
            store,
            retrieved_log,
        )
        # log the grouped component-level retrieval (incl. standards) for the export sheet
        for group, matches in retrieve_groups(store, inputs).items():
            for rank, match in enumerate(matches, start=1):
                meta = match.get("metadata", {})
                retrieved_log.append(
                    {
                        "Query ID": f"Q-COMPONENT-{group.upper()}",
                        "Generation Type": f"Component context ({group})",
                        "Source Rank": rank,
                        "Similarity Score": match["similarity"],
                        "Document Type": meta.get("document_type", ""),
                        "Source File": meta.get("file_name", ""),
                        "Source Sheet": meta.get("sheet_name", ""),
                        "Source Row": meta.get("row_number", ""),
                        "Source Chunk ID": match.get("chunk_id", ""),
                        "Chunk Preview": str(match.get("chunk_text", ""))[:180],
                        "Used In Output": "Reference",
                    }
                )
        retrieved_df = pd.DataFrame(retrieved_log, columns=RETRIEVED_SOURCES_COLUMNS)
        return dfmea_df, dvp_df, lessons_df, retrieved_df
    except Exception:
        return dfmea_df, dvp_df, lessons_df, empty_log


def apply_llm_enrichment(
    dfmea_df: pd.DataFrame,
    dvp_df: pd.DataFrame,
    inputs: dict[str, str],
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Optional roadmap Phase 7-8: LLM generation from retrieved context.

    Off unless RAG_LLM_PROVIDER is configured. All LLM rows are schema-validated,
    citation-checked against real chunk IDs, appended as clearly-labeled
    'LLM + RAG' suggestions, and require engineer review.
    """
    store = get_rag_store()
    retrieved = retrieve_groups(store, inputs)
    known_chunks: dict[str, dict[str, Any]] = {}
    for group in retrieved.values():
        for match in group:
            known_chunks[match["chunk_id"]] = match
    prompt = build_rag_prompt(
        inputs,
        retrieved,
        baseline_failure_modes=[str(x) for x in dfmea_df.get("Potential Failure Mode", pd.Series(dtype=str))],
        baseline_tests=[str(x) for x in dvp_df.get("Recommended Validation Test", pd.Series(dtype=str))],
    )
    llm_dfmea, llm_dvpr, notes = generate_llm_enrichment(prompt, set(known_chunks))

    def _source_fields(chunk_ids: list[str]) -> dict[str, Any]:
        if not chunk_ids:
            return {
                "Source Evidence": "LLM suggestion without retrieved source - requires engineer verification",
                "AI Confidence": 0.5,
                "AI Confidence Score": 0.5,
            }
        match = known_chunks[chunk_ids[0]]
        meta = match.get("metadata", {})
        conf = _confidence_from_similarity(match["similarity"])
        return {
            "Source Evidence": str(match.get("chunk_text", ""))[:180],
            "Source File": meta.get("file_name", ""),
            "Source Sheet": meta.get("sheet_name", ""),
            "Source Row": meta.get("row_number", ""),
            "Source Chunk ID": match.get("chunk_id", ""),
            "Source Strength": meta.get("source_strength", "Unknown"),
            "Citation Count": len(chunk_ids),
            "Future RAG Citation": citation_label(match),
            "AI Confidence": conf,
            "AI Confidence Score": conf,
        }

    new_dfmea_rows = []
    for i, row in enumerate(llm_dfmea, start=1):
        base = dfmea_row(
            inputs,
            str(row["failure_mode"]),
            str(row["effect"]),
            int(row["severity"]),
            str(row["cause"]),
            int(row["occurrence"]),
            str(row.get("prevention_control", "Engineer to define prevention control")),
            str(row.get("detection_control", "Engineer to define detection control")),
            int(row["detection"]),
            str(row["recommended_action"]),
            str(row.get("owner", "BIW")),
            str(row.get("rationale", "LLM-suggested from retrieved context")),
        )
        base.update(
            {
                "Failure Mode ID": f"FM-AI-{i:02d}",
                "Action ID": f"ACT-AI-{i:02d}",
                "AI Suggestion ID": f"AI-LLM-DFMEA-{i:03d}",
                "Source Type": "LLM + RAG",
                "Review Status": "Under Review - LLM suggestion",
                "Human Review Required": "Yes",
                "Reason for Human Review": "LLM-generated suggestion",
                "Change Log": "Generated by LLM from retrieved context; pending engineer review.",
            }
        )
        base.update(_source_fields(row.get("source_chunk_ids", [])))
        new_dfmea_rows.append(base)

    new_dvp_rows = []
    if not dfmea_df.empty or new_dfmea_rows:
        lookup_df = pd.concat([dfmea_df, pd.DataFrame(new_dfmea_rows, columns=DFMEA_COLUMNS)]) if new_dfmea_rows else dfmea_df
        for i, row in enumerate(llm_dvpr, start=1):
            linked = str(row["linked_failure_mode"])
            matches = lookup_df[lookup_df["Potential Failure Mode"].astype(str).str.lower() == linked.lower()]
            anchor = matches.iloc[0] if not matches.empty else lookup_df.iloc[0]
            base = dvp_row(
                anchor,
                str(row["test"]),
                str(row.get("validation_type", "Physical")),
                str(row["objective"]),
                str(row["acceptance_criteria"]),
                str(row.get("team", "BIW / Validation")),
                str(row.get("rationale", "LLM-suggested from retrieved context")),
                validation_status="Proposed",
                test_id=f"TEST-AI-{i:02d}",
            )
            base.update(
                {
                    "AI Suggestion ID": f"AI-LLM-DVPR-{i:03d}",
                    "Source Type": "LLM + RAG",
                    "Review Status": "Under Review - LLM suggestion",
                    "Human Review Required": "Yes",
                    "Reason for Human Review": "LLM-generated suggestion",
                    "Change Log": "Generated by LLM from retrieved context; pending engineer review.",
                }
            )
            base.update(_source_fields(row.get("source_chunk_ids", [])))
            new_dvp_rows.append(base)

    if new_dfmea_rows:
        dfmea_df = pd.concat([dfmea_df, pd.DataFrame(new_dfmea_rows, columns=DFMEA_COLUMNS)], ignore_index=True)
    if new_dvp_rows:
        dvp_df = pd.concat([dvp_df, pd.DataFrame(new_dvp_rows, columns=DVP_COLUMNS)], ignore_index=True)
    return dfmea_df, dvp_df, notes


def augment_gap_analysis(
    gap_df: pd.DataFrame,
    dfmea_df: pd.DataFrame,
    dvp_df: pd.DataFrame,
    retrieved_df: pd.DataFrame,
) -> pd.DataFrame:
    """Add source-aware roadmap gap types on top of validation-coverage gaps."""
    rows: list[dict[str, Any]] = []
    next_id = len(gap_df) + 1

    def add(gap_type: str, description: str, fix: str, severity: int = 0, failure_mode: str = "", fm_id: str = "", priority: str = "Medium") -> None:
        nonlocal next_id
        rows.append(
            {
                "Gap ID": f"GAP-{next_id:03d}",
                "Failure Mode ID": fm_id,
                "Failure Mode": failure_mode,
                "Severity": severity,
                "Gap Type": gap_type,
                "Gap Description": description,
                "Recommended Fix": fix,
                "Priority": priority,
                "Status": "Open",
                "Gap Status": "Open",
                "Engineer Decision": "Pending",
                "Gap Closure Status": "Open",
                "Final Gap Closure Status": "Open",
                "Approval Status": "Draft",
            }
        )
        next_id += 1

    if not dvp_df.empty and not dfmea_df.empty:
        known_fm_ids = set(dfmea_df["Failure Mode ID"].astype(str))
        for _, test in dvp_df.iterrows():
            linked = str(test.get("Linked Failure Mode ID", "")).strip()
            if linked and linked not in known_fm_ids:
                add(
                    "DVP&R test without linked DFMEA risk",
                    f"Test {test.get('Test ID', '?')} ({test.get('Recommended Validation Test', '')}) links to unknown failure mode ID {linked}.",
                    "Re-link the test to a valid DFMEA failure mode or remove it.",
                )
        dfmea_reqs = set(dfmea_df["Requirement ID"].astype(str)) - {""}
        dvp_reqs = set(dvp_df["Requirement ID"].astype(str)) - {""}
        for req in sorted(dfmea_reqs - dvp_reqs):
            add(
                "Requirement without linked validation",
                f"Requirement {req} appears in the DFMEA but has no DVP&R test linked to it.",
                "Add a validation test covering this requirement or document a waiver.",
                priority="High",
            )

    if not dfmea_df.empty and "Source Chunk ID" in dfmea_df.columns:
        high_risk = dfmea_df[(dfmea_df["Severity"] >= 8) & (dfmea_df["Source Chunk ID"].astype(str) == "")]
        for _, row in high_risk.iterrows():
            add(
                "No source found for high-risk recommendation",
                f"High-severity failure mode '{row['Potential Failure Mode']}' (S={row['Severity']}) has no retrieved source evidence.",
                "Verify against historical DFMEAs and standards manually, or expand the knowledge base.",
                severity=int(row["Severity"]),
                failure_mode=str(row["Potential Failure Mode"]),
                fm_id=str(row["Failure Mode ID"]),
                priority="High",
            )

    if not retrieved_df.empty and not dfmea_df.empty:
        used = set(dfmea_df.get("Source Chunk ID", pd.Series(dtype=str)).astype(str)) | set(
            dvp_df.get("Source Chunk ID", pd.Series(dtype=str)).astype(str)
        )
        strong_context = retrieved_df[
            (retrieved_df["Generation Type"].astype(str).str.startswith("Component context"))
            & (retrieved_df["Similarity Score"].astype(float) >= 0.5)
        ]
        for _, src in strong_context.iterrows():
            if str(src["Source Chunk ID"]) not in used:
                add(
                    "RAG source found but not used",
                    f"Strong knowledge-base match ({src['Source File']} row {src['Source Row']}, similarity {src['Similarity Score']}) was retrieved but no generated row cites it.",
                    "Review this source for a failure mode or validation item the draft may be missing.",
                    priority="Medium",
                )

    if not rows:
        return gap_df
    return pd.concat([gap_df, pd.DataFrame(rows, columns=GAP_COLUMNS)], ignore_index=True)


def kb_summary_dataframe() -> pd.DataFrame:
    rows: list[tuple[str, str]] = []
    if RAG_IMPORT_ERROR:
        rows.append(("Status", f"RAG layer unavailable: {RAG_IMPORT_ERROR}"))
    else:
        try:
            stats = get_rag_store().get_collection_stats()
            rows += [
                ("Indexed Documents", str(stats["documents"])),
                ("Indexed Chunks", str(stats["chunks"])),
                ("Document Types", ", ".join(f"{k} ({v})" for k, v in sorted(stats["document_types"].items())) or "None"),
                ("Source Strength Breakdown", ", ".join(f"{k} ({v})" for k, v in sorted(stats["source_strengths"].items())) or "None"),
                ("Last Indexed", stats["last_indexed"] or "Never"),
                ("Vector Database Path", stats["path"]),
                ("Embedding Model", stats["embedding_model"]),
            ]
        except Exception as exc:
            rows.append(("Status", f"Knowledge base not readable: {exc}"))
    rows.append(("Disclaimer", RAG_DISCLAIMER))
    return pd.DataFrame(rows, columns=["Field", "Value"])


def init_state() -> None:
    for field in COMPONENT_FIELDS:
        st.session_state.setdefault(field, "")
    st.session_state.setdefault("selected_part_label", DEFAULT_PART_LABEL)
    st.session_state.setdefault("loaded_part_label", "")
    st.session_state.setdefault("demo_gap_mode", True)
    st.session_state.setdefault("dfmea_df", empty_df(DFMEA_COLUMNS))
    st.session_state.setdefault("dvp_df", empty_df(DVP_COLUMNS))
    st.session_state.setdefault("trace_df", empty_df(TRACE_COLUMNS))
    st.session_state.setdefault("gap_df", empty_df(GAP_COLUMNS))
    st.session_state.setdefault("lessons_df", empty_df(LESSON_COLUMNS))
    st.session_state.setdefault("pdiag_df", empty_df(P_DIAGRAM_COLUMNS))
    st.session_state.setdefault("retrieved_sources_df", empty_df(RETRIEVED_SOURCES_COLUMNS))
    st.session_state.setdefault("llm_enrich_mode", False)
    st.session_state.setdefault("llm_notes", [])
    st.session_state.setdefault("last_generated", False)
    # Auto-seed the knowledge base so RAG works out of the box (incl. ephemeral HF Spaces).
    if not RAG_IMPORT_ERROR and not st.session_state.get("kb_seed_attempted"):
        st.session_state["kb_seed_attempted"] = True
        try:
            store = get_rag_store()
            if store.get_collection_stats()["chunks"] == 0:
                seed_knowledge_base(store)
        except Exception:
            pass


def component_inputs() -> dict[str, str]:
    return {field: str(st.session_state.get(field, "")).strip() for field in COMPONENT_FIELDS}


def generation_settings() -> dict[str, Any]:
    return {field: st.session_state.get(field) for field in GENERATION_SETTINGS}


def make_context(inputs: dict[str, str]) -> RuleContext:
    text = " ".join(str(value).lower() for value in inputs.values())
    return RuleContext(inputs=inputs, text=text)


def item_name(inputs: dict[str, str]) -> str:
    return inputs.get("component_name") or "BIW sheet metal component"


def primary_function(inputs: dict[str, str]) -> str:
    return inputs.get("primary_function") or "Meet intended structural, dimensional, and durability requirements"


def rpn(severity: int, occurrence: int, detection: int) -> int:
    return int(severity) * int(occurrence) * int(detection)


def action_priority(severity: int, occurrence: int, rpn_value: int) -> str:
    if severity >= 9:
        return "High"
    if severity >= 8 and occurrence >= 4:
        return "High"
    if severity >= 7 and rpn_value >= 120:
        return "High"
    if rpn_value >= 80:
        return "Medium"
    return "Low"


def aiag_vda_action_priority(severity: int, occurrence: int, detection: int) -> str:
    """Condensed implementation of the AIAG-VDA 2019 Action Priority (AP) logic.

    The official handbook uses a 1000-cell S/O/D lookup table. This function reproduces
    the table's intent in banded form: AP is driven first by Severity, then Occurrence,
    then Detection - it is NOT an RPN threshold. Production use must adopt the company's
    approved AP table verbatim.
    """
    s, o, d = int(severity), int(occurrence), int(detection)
    if s >= 9:
        if o >= 4 or (o >= 2 and d >= 5):
            return "H"
        if o >= 2 or d >= 7:
            return "M"
        return "L"
    if s >= 7:
        if o >= 6 or (o >= 4 and d >= 5):
            return "H"
        if o >= 4 or (o >= 2 and d >= 5) or d >= 7:
            return "M"
        return "L"
    if s >= 4:
        if o >= 8 or (o >= 6 and d >= 7):
            return "H"
        if o >= 4 and d >= 5:
            return "M"
        return "L"
    return "L"


def special_characteristic(severity: int, occurrence: int) -> str:
    """Flag potential special characteristics for engineering review.

    Severity 9-10 -> potential critical characteristic (YC): safety or regulatory effect.
    Severity 5-8 with elevated occurrence -> potential significant characteristic (SC).
    Final designation always belongs to the responsible engineer and program SC process.
    """
    if severity >= 9:
        return "YC (Potential Critical)"
    if severity >= 5 and occurrence >= 4:
        return "SC (Potential Significant)"
    return ""


def end_effect_for_category(category: str) -> str:
    mapping = {
        "Crash / Safety": "Degraded occupant protection or regulatory non-compliance at vehicle level",
        "Durability": "Reduced vehicle service life; potential warranty repair and customer dissatisfaction",
        "Joining / Durability": "Structural joint degradation over vehicle life; potential noise and warranty claim",
        "Manufacturing Quality": "Inconsistent body structure quality reaching the customer; rework and cost impact",
        "Corrosion": "Visible corrosion or perforation in the field; durability and perceived-quality impact",
        "Dimensional": "Poor fit-and-finish, wind noise, or water leak perceived by the customer",
        "Stamping / Formability": "Part quality escape or launch delay; potential durability margin reduction",
        "NVH": "Buzz, squeak, rattle, or booming perceived by the customer during operation",
        "Attachment / Durability": "Loss of component attachment function in service; noise and safety margin impact",
        "Serviceability": "Extended repair time and cost of ownership impact for the customer",
        "Assembly": "Assembly line disruption, rework, or quality escape to the customer",
    }
    return mapping.get(category, "Reduced vehicle-level performance, quality, or customer satisfaction")


def residual_risk_level(rpn_value: int) -> str:
    if rpn_value >= 100:
        return "High"
    if rpn_value >= 50:
        return "Medium"
    return "Low"


def risk_category_for_failure(failure_mode: str) -> str:
    mode = failure_mode.lower()
    if "crash" in mode or "load-path" in mode or "load path" in mode:
        return "Crash / Safety"
    if "spot weld fatigue" in mode:
        return "Joining / Durability"
    if "weld" in mode or "nugget" in mode:
        return "Manufacturing Quality"
    if "fatigue" in mode or "crack" in mode:
        return "Durability"
    if "corrosion" in mode or "e-coat" in mode or "flange" in mode:
        return "Corrosion"
    if "dimensional" in mode or "variation" in mode or "fit" in mode:
        return "Dimensional"
    if "forming" in mode or "thinning" in mode or "wrinkling" in mode:
        return "Stamping / Formability"
    if "nvh" in mode or "vibration" in mode or "stiffness" in mode:
        return "NVH"
    if "pull-out" in mode or "fastener" in mode or "bracket" in mode:
        return "Attachment / Durability"
    if "service" in mode:
        return "Serviceability"
    return "Assembly"


def function_id_for_category(category: str) -> str:
    mapping = {
        "Crash / Safety": "F-001",
        "Durability": "F-002",
        "Joining / Durability": "F-002",
        "Dimensional": "F-003",
        "Corrosion": "F-004",
        "Stamping / Formability": "F-005",
        "Manufacturing Quality": "F-006",
        "NVH": "F-007",
        "Attachment / Durability": "F-008",
        "Serviceability": "F-009",
    }
    return mapping.get(category, "F-010")


def requirement_id_for_category(category: str) -> str:
    mapping = {
        "Crash / Safety": "REQ-CRASH-001",
        "Durability": "REQ-DUR-001",
        "Joining / Durability": "REQ-JOIN-001",
        "Manufacturing Quality": "REQ-JOIN-001",
        "Corrosion": "REQ-COR-001",
        "Dimensional": "REQ-DIM-001",
        "Stamping / Formability": "REQ-MFG-001",
        "NVH": "REQ-NVH-001",
        "Attachment / Durability": "REQ-ATTACH-001",
        "Serviceability": "REQ-SERVICE-001",
    }
    return mapping.get(category, "REQ-GEN-001")


def revised_ratings(severity: int, occurrence: int, detection: int) -> tuple[int, int, int]:
    revised_severity = severity
    revised_occurrence = max(1, occurrence - 1)
    revised_detection = max(1, detection - 1)
    return revised_severity, revised_occurrence, revised_detection


def human_review_reason(
    severity: int,
    ap: str,
    confidence: float,
    source_type: str,
    coverage_status: str | None = None,
) -> str:
    reasons: list[str] = []
    if severity >= 8:
        reasons.append("Severity >= 8")
    if ap == "High":
        reasons.append("Action Priority = High")
    if confidence < 0.75:
        reasons.append("AI Confidence Score < 0.75")
    if source_type == "Synthetic MVP Rule":
        reasons.append("Source Type = Synthetic MVP Rule")
    if coverage_status in {"Gap", "Partial"}:
        reasons.append(f"Coverage Status = {coverage_status}")
    return "; ".join(reasons) if reasons else "Standard engineer review"


def review_required(reason: str) -> str:
    return "Yes" if reason and reason != "Standard engineer review" else "No"


def dfmea_row(
    inputs: dict[str, str],
    failure_mode: str,
    effect: str,
    severity: int,
    cause: str,
    occurrence: int,
    prevention: str,
    detection_control: str,
    detection: int,
    action: str,
    owner: str,
    rationale: str,
) -> dict[str, Any]:
    category = risk_category_for_failure(failure_mode)
    initial_rpn = rpn(severity, occurrence, detection)
    revised_severity, revised_occurrence, revised_detection = revised_ratings(severity, occurrence, detection)
    revised_rpn = rpn(revised_severity, revised_occurrence, revised_detection)
    reduction = round((initial_rpn - revised_rpn) / initial_rpn, 3) if initial_rpn else 0
    ap = action_priority(severity, occurrence, initial_rpn)
    ap_vda = aiag_vda_action_priority(severity, occurrence, detection)
    sc_flag = special_characteristic(severity, occurrence)
    end_effect = end_effect_for_category(category)
    confidence = 0.7
    source_type = "Synthetic MVP Rule"
    reason = human_review_reason(severity, ap, confidence, source_type)
    final_text = f"{failure_mode}: {action}"

    return {
        "Failure Mode ID": "",
        "Function ID": function_id_for_category(category),
        "Requirement ID": requirement_id_for_category(category),
        "Action ID": "",
        "Item": item_name(inputs),
        "Function": primary_function(inputs),
        "Risk Category": category,
        "Potential Failure Mode": failure_mode,
        "Potential Effect of Failure": effect,
        "End Effect (Vehicle Level)": end_effect,
        "Special Characteristic": sc_flag,
        "Potential Cause / Mechanism": cause,
        "Initial Severity": severity,
        "Initial Occurrence": occurrence,
        "Initial Detection": detection,
        "Initial RPN": initial_rpn,
        "Severity": severity,
        "Occurrence": occurrence,
        "Detection": detection,
        "RPN": initial_rpn,
        "Action Priority": ap,
        "AP (AIAG-VDA)": ap_vda,
        "Prevention Control": prevention,
        "Detection Control": detection_control,
        "Recommended Action": action,
        "Revised Severity": revised_severity,
        "Revised Occurrence": revised_occurrence,
        "Revised Detection": revised_detection,
        "Revised RPN": revised_rpn,
        "RPN Reduction %": reduction,
        "Residual Risk Level": residual_risk_level(revised_rpn),
        "Action Owner": owner,
        "Responsible Team": owner,
        "Target Completion Date": "TBD",
        "Action Status": "Open",
        "Closure Evidence": "TBD",
        "Closure Notes": "Engineer Review Required",
        "Notes / Rationale": rationale,
        "AI Suggestion ID": "",
        "AI Confidence Score": confidence,
        "Source Type": source_type,
        "Source Strength": "Low",
        "Citation Count": 0,
        "Human Review Required": review_required(reason),
        "Reason for Human Review": reason,
        "Future RAG Citation": "Not connected",
        "Source Evidence": "",
        "Source File": "",
        "Source Sheet": "",
        "Source Row": "",
        "Source Chunk ID": "",
        "AI Confidence": confidence,
        "Review Status": "Draft",
        "Engineer Decision": "Pending",
        "Rejection Reason": "",
        "Final Approved Text": final_text,
        "Reviewed By": "TBD",
        "Review Date": "",
        "Approval Status": "Draft",
        "Approval Notes": "",
        "Change Log": "Initial rules-based draft generated by prototype.",
    }


def finalize_dfmea(rows: list[dict[str, Any]]) -> pd.DataFrame:
    def sort_key(row: dict[str, Any]) -> int:
        failure_mode = str(row.get("Potential Failure Mode", "")).lower()
        if "bracket or fastener pull-out" in failure_mode:
            return 85
        if "material thinning" in failure_mode or "wrinkling" in failure_mode:
            return 90
        return 50

    rows = sorted(enumerate(rows), key=lambda item: (sort_key(item[1]), item[0]))
    rows = [row for _, row in rows]
    for index, row in enumerate(rows, start=1):
        row["Failure Mode ID"] = f"FM-{index:03d}"
        row["Action ID"] = f"ACT-{index:03d}"
        row["AI Suggestion ID"] = f"AI-DFMEA-{index:03d}"
    return pd.DataFrame(rows, columns=DFMEA_COLUMNS)


def generate_dfmea(inputs: dict[str, str]) -> pd.DataFrame:
    ctx = make_context(inputs)
    rows: list[dict[str, Any]] = []

    rows.append(
        dfmea_row(
            inputs,
            "Dimensional variation during body build",
            "Poor fit-up, downstream assembly issues, or reduced dimensional quality",
            6,
            "Stack-up sensitivity, weak locating scheme, forming variation, or fixture variation",
            4,
            "GD&T review, locating strategy review, forming feasibility review",
            "Dimensional build study, fixture buyoff, body measurement report",
            5,
            "Review datum strategy, tolerance stack, locating holes, and build sequence",
            "BIW / Dimensional Engineering",
            "Baseline dimensional risk for stamped and welded BIW structures.",
        )
    )

    if ctx.has_any(["fatigue", "durability", "road load", "torsional"]):
        rows.append(
            dfmea_row(
                inputs,
                "Fatigue crack at formed radius",
                "Reduced durability performance, local stiffness loss, or potential crack propagation",
                8,
                "High local stress concentration, small radius, material thinning, or load path discontinuity",
                4,
                "CAE durability review, forming radius guideline, gauge and bead strategy review",
                "Physical durability test, teardown inspection, CAE correlation review",
                5,
                "Review local radius, gauge, bead strategy, and stress concentration",
                "BIW / CAE",
                "Triggered by durability, fatigue, or torsional load concerns.",
            )
        )

    if ctx.has_any(["spot weld", "spotweld", "weld"]):
        rows.append(
            dfmea_row(
                inputs,
                "Spot weld fatigue",
                "Joint separation, reduced load transfer, or durability degradation",
                8,
                "Insufficient weld pitch, thin flange, poor load distribution, or high cyclic load",
                4,
                "Weld layout review, load-path review, weld standard compliance",
                "Weld teardown after durability exposure, coupon fatigue review",
                4,
                "Review weld pitch, flange width, joint load path, and local stiffness",
                "BIW / CAE / Validation",
                "Triggered by welded joining strategy and durability-sensitive BIW joints.",
            )
        )
        rows.append(
            dfmea_row(
                inputs,
                "Spot weld quality or nugget nonconformance",
                "Reduced joint strength, manufacturing rework, or inconsistent body structure quality",
                7,
                "Poor gun access, process window variation, shunting, contamination, or flange stack variation",
                4,
                "Weld access review, process feasibility review, weld standard compliance",
                "Nugget inspection, peel test, process audit, weld destruct testing",
                4,
                "Review weld gun access, nugget size, shunting risk, and process controls",
                "BIW / Manufacturing / Quality",
                "Separated from weld fatigue so validation recommendations stay focused.",
            )
        )

    if ctx.has_any(["adhesive", "bond"]):
        rows.append(
            dfmea_row(
                inputs,
                "Adhesive bond degradation",
                "Reduced joint stiffness, moisture ingress, or reduced load transfer",
                7,
                "Poor surface preparation, adhesive gap variation, cure variation, or environmental aging",
                3,
                "Bond design review, surface prep controls, process window review",
                "Lap shear test, environmental aging, durability test, process audit",
                5,
                "Review adhesive bead path, bond gap, surface prep, and cure process controls",
                "BIW / Materials / Manufacturing",
                "Triggered by structural adhesive or bonded interface.",
            )
        )

    if ctx.has_any(["corrosion", "underbody", "water", "salt", "flange", "overlap", "splash"]):
        rows.append(
            dfmea_row(
                inputs,
                "Corrosion at flange overlap",
                "Long-term durability degradation, perforation risk, or reduced joint integrity",
                7,
                "Water trap, poor drainage, poor sealing, limited e-coat access, or galvanic sensitivity",
                4,
                "Drainage review, sealer strategy review, flange geometry review",
                "Corrosion cycle test, e-coat coverage review, sectioning inspection",
                5,
                "Add drain path, improve sealer path, review e-coat access, and avoid water traps",
                "Materials / Corrosion / BIW",
                "Triggered by corrosion exposure, underbody area, or overlapped flanges.",
            )
        )
        rows.append(
            dfmea_row(
                inputs,
                "Poor e-coat coverage in closed section or flange",
                "Reduced corrosion protection and increased long-term field risk",
                7,
                "Restricted coating flow, closed geometry, tight flange, or missing drain and vent path",
                3,
                "E-coat access review, drain and vent design review",
                "E-coat coverage study, sectioning inspection, corrosion validation",
                6,
                "Review drain holes, venting, section openness, and flange geometry",
                "Materials / Corrosion",
                "Coverage risk is common in closed or partially closed BIW sections.",
            )
        )

    if ctx.has_any(["crash", "impact", "load path", "load-path", "rear impact", "front impact", "roof crush"]):
        rows.append(
            dfmea_row(
                inputs,
                "Crash load-path deformation or instability",
                "Reduced crash energy management or unintended load transfer",
                10,
                "Section instability, weak reinforcement continuity, poor joint layout, or local buckling",
                3,
                "Crash CAE review, load-path design review, section stability review",
                "Crash CAE, physical crash validation, post-test teardown",
                4,
                "Add CAE correlation and physical crash validation to confirm load-path behavior",
                "BIW / Safety / CAE",
                "Triggered by crash, impact, or roof crush load cases.",
            )
        )

    if ctx.has_any(["n vh", "nvh", "buzz", "squeak", "rattle", "stiffness", "modal"]):
        rows.append(
            dfmea_row(
                inputs,
                "NVH vibration due to insufficient local stiffness",
                "Buzz, squeak, rattle, customer dissatisfaction, or perceived quality issue",
                5,
                "Low local stiffness, unsupported panel span, joint compliance, or bracket resonance",
                4,
                "Modal review, stiffness target review, joint strategy review",
                "Modal analysis, body stiffness test, road noise evaluation",
                5,
                "Review local stiffness, bead strategy, attachment points, and resonance risk",
                "BIW / NVH",
                "Triggered by NVH, stiffness, or vibration concerns.",
            )
        )

    if ctx.has_any(["forming", "stamp", "stamped", "wrinkle", "wrinkling", "thinning", "split"]):
        rows.append(
            dfmea_row(
                inputs,
                "Material thinning or wrinkling during forming",
                "Reduced durability margin, poor fit, scrap, rework, or part quality issue",
                6,
                "Aggressive draw depth, tight radius, material flow restriction, or trim sensitivity",
                4,
                "Forming simulation, die feasibility review, material selection review",
                "Tryout inspection, thickness measurement, forming simulation correlation",
                5,
                "Review draw strategy, radius, bead placement, trim line, and local thinning",
                "Stamping / Manufacturing / BIW",
                "Triggered by stamped sheet metal manufacturing process.",
            )
        )

    if ctx.has_any(["bracket", "mount", "fastener", "bolt"]):
        rows.append(
            dfmea_row(
                inputs,
                "Bracket or fastener pull-out under load",
                "Loss of attachment, local deformation, noise, or durability degradation",
                8,
                "Insufficient local reinforcement, poor load spread, thin material, or edge distance issue",
                3,
                "Joint design review, local reinforcement review, fastener standard review",
                "Pull-out test, torque audit, durability test, teardown inspection",
                5,
                "Review edge distance, washer strategy, local gauge, embossments, and reinforcement need",
                "BIW / Validation / Manufacturing",
                "Triggered by bracket, mount, or fastener interface.",
            )
        )

    if ctx.has_any(["service", "serviceability"]):
        rows.append(
            dfmea_row(
                inputs,
                "Service access constraint at attachment interface",
                "Difficult service operation, repair delay, or risk of incorrect reassembly",
                5,
                "Insufficient tool clearance, blocked fastener access, or unclear service path",
                3,
                "Serviceability review, tool clearance review, assembly sequence review",
                "Service mockup review, digital human factors check",
                5,
                "Review service tool access, fastener approach, and removal sequence",
                "BIW / Service / Manufacturing",
                "Triggered by serviceability concern.",
            )
        )

    return finalize_dfmea(rows)


def generate_p_diagram(inputs: dict[str, str]) -> pd.DataFrame:
    """Generate a Parameter (P) Diagram supporting AIAG-VDA function analysis.

    The P-Diagram documents the ideal function, input signal, control factors,
    the five standard noise factor categories, and error states. It is the bridge
    between function analysis and failure analysis in the 7-step approach.
    """
    ctx = make_context(inputs)
    rows: list[tuple[str, str, str]] = []

    rows.append(("Ideal Function", "Function", primary_function(inputs)))
    rows.append(("Input Signal", "Function", inputs.get("load_cases") or "Structural loads transferred from interfacing body structure"))
    rows.append(("Intended Output", "Function", "Loads transferred and geometry maintained per requirement over vehicle life"))

    control_factors = [
        f"Material selection: {inputs.get('material') or 'TBD'}",
        f"Gauge / thickness: {inputs.get('thickness') or 'TBD'}",
        f"Joining strategy: {inputs.get('joining_method') or 'TBD'}",
        f"Manufacturing process: {inputs.get('manufacturing_process') or 'TBD'}",
        "Section geometry, bead strategy, and local reinforcement",
        "Datum and locating strategy (GD&T)",
    ]
    for factor in control_factors:
        rows.append(("Control Factor", "Design Parameter", factor))

    piece_noise = ["Stamping thickness and springback variation", "fixture and locating variation"]
    if ctx.has_any(["weld"]):
        piece_noise.insert(1, "weld nugget size and pitch variation")
    if ctx.has_any(["adhesive", "bond"]):
        piece_noise.append("adhesive bead placement and bond gap variation")
    if ctx.has_any(["fastener", "bolt", "bracket", "mount"]):
        piece_noise.append("fastener torque and hole position variation")
    rows.append(("Noise: Piece-to-Piece Variation", "Noise Factor", ", ".join(piece_noise).capitalize()))

    time_noise = ["Fatigue damage accumulation", "joint relaxation"]
    if ctx.has_any(["corrosion", "underbody", "salt", "splash"]):
        time_noise.insert(1, "corrosion progression")
    if ctx.has_any(["adhesive", "bond"]):
        time_noise.append("adhesive aging and environmental degradation")
    rows.append(("Noise: Changes Over Time", "Noise Factor", ", ".join(time_noise).capitalize()))
    rows.append(("Noise: Customer Usage", "Noise Factor", "Severe road inputs, overload events, towing/payload extremes, high-mileage duty cycles"))
    rows.append(("Noise: External Environment", "Noise Factor", inputs.get("environmental_exposure") or "Temperature cycling, humidity, road salt, stone impingement"))
    rows.append(("Noise: System Interaction", "Noise Factor", inputs.get("interfaces") or "Load and tolerance interaction with mating body structure"))

    error_states = [
        "Loss or degradation of the primary load transfer function (intended output not achieved)",
        "Dimensional variation beyond tolerance during body build",
    ]
    if ctx.has_any(["fatigue", "durability", "torsional"]):
        error_states.append("Fatigue crack initiation at stress concentration")
    if ctx.has_any(["weld", "spot weld"]):
        error_states.append("Weld joint fatigue or nugget nonconformance")
    if ctx.has_any(["adhesive", "bond"]):
        error_states.append("Adhesive bond degradation or disbond")
    if ctx.has_any(["corrosion", "underbody", "salt", "flange"]):
        error_states.append("Corrosion initiation at flange or closed section")
    if ctx.has_any(["crash", "impact", "load path", "load-path"]):
        error_states.append("Unstable crash load-path deformation")
    if ctx.has_any(["nvh", "stiffness", "rattle", "buzz"]):
        error_states.append("NVH degradation from insufficient local stiffness")
    if ctx.has_any(["bracket", "mount", "fastener", "bolt"]):
        error_states.append("Attachment pull-out or loosening under load")
    for state in error_states:
        rows.append(("Error State", "Failure Mode Link", state))

    rows.append(("Methodology", "Reference", METHODOLOGY_NOTE))
    return pd.DataFrame(rows, columns=P_DIAGRAM_COLUMNS)


def fmea_header_dataframe(inputs: dict[str, str]) -> pd.DataFrame:
    """AIAG-VDA planning & preparation: FMEA header / identification block."""
    component = inputs.get("component_name") or "BIW Component"
    fmea_id = "DFMEA-BIW-" + "".join(word[0] for word in component.split()[:4]).upper() + "-001"
    rows = [
        ("FMEA ID", fmea_id),
        ("Subject", component),
        ("Vehicle Area / System", inputs.get("vehicle_area", "")),
        ("Program Phase", inputs.get("program_phase") or "TBD"),
        ("Design Responsibility", inputs.get("engineer_name") or "TBD"),
        ("Core Team", "BIW design, CAE, validation, manufacturing, quality, materials (to be confirmed)"),
        ("FMEA Start Date", date.today().isoformat()),
        ("FMEA Revision Date", date.today().isoformat()),
        ("Revision", TOOL_VERSION + " draft"),
        ("Methodology", METHODOLOGY_NOTE),
        ("Confidentiality", "Synthetic prototype data - no confidential program content"),
    ]
    return pd.DataFrame(rows, columns=["Field", "Value"])


def validation_profile(test: str, validation_type: str) -> dict[str, str]:
    text = f"{test} {validation_type}".lower()
    if "crash" in text:
        return {
            "Validation Level": "Subsystem / Vehicle",
            "Test Method / Procedure": "Crash CAE correlation and physical crash validation procedure",
            "Test Standard Reference": "FMVSS / internal crash procedure CR-BIW-001 (synthetic ref)",
            "Test Stage": "DV",
            "Build Phase": "Alpha",
            "Sample Size": "1 CAE model + 1 vehicle or subsystem build",
            "Test Duration / Cycles": "One crash event with pre and post-test teardown",
            "Test Conditions": "Representative regulated or internal crash load case",
        }
    if "durability" in text or "fatigue" in text:
        return {
            "Validation Level": "Component",
            "Test Method / Procedure": "CAE durability review plus physical durability schedule",
            "Test Standard Reference": "TS-DUR-BIW-014 durability schedule (synthetic ref)",
            "Test Stage": "DV",
            "Build Phase": "Beta",
            "Sample Size": "3 components minimum",
            "Test Duration / Cycles": "Program durability target cycles",
            "Test Conditions": "Representative road load and body torsion inputs",
        }
    if "weld" in text or "nugget" in text:
        return {
            "Validation Level": "Coupon",
            "Test Method / Procedure": "Weld destruct, peel, nugget, and teardown inspection",
            "Test Standard Reference": "WS-JOIN-003 weld quality standard (synthetic ref)",
            "Test Stage": "DV / PV",
            "Build Phase": "Alpha",
            "Sample Size": "10 weld coupons or representative body joints",
            "Test Duration / Cycles": "Per weld validation plan",
            "Test Conditions": "Nominal and boundary process window",
        }
    if "corrosion" in text or "e-coat" in text:
        return {
            "Validation Level": "Component",
            "Test Method / Procedure": "Corrosion cycle test and coating section review",
            "Test Standard Reference": "TS-COR-021 cyclic corrosion procedure (synthetic ref)",
            "Test Stage": "DV / PV",
            "Build Phase": "Beta",
            "Sample Size": "3 components minimum",
            "Test Duration / Cycles": "Program corrosion cycle target",
            "Test Conditions": "Underbody splash, salt, and humidity exposure",
        }
    if "dimensional" in text:
        return {
            "Validation Level": "Subsystem",
            "Test Method / Procedure": "Body build dimensional measurement study",
            "Test Standard Reference": "DIM-BUILD-007 body dimensional procedure (synthetic ref)",
            "Test Stage": "PV",
            "Build Phase": "Alpha",
            "Sample Size": "10 builds or measurement events",
            "Test Duration / Cycles": "Build event duration",
            "Test Conditions": "Representative fixture, locator, and joining process",
        }
    if "forming" in text or "thickness" in text:
        return {
            "Validation Level": "Component",
            "Test Method / Procedure": "Forming simulation and tryout thickness inspection",
            "Test Standard Reference": "FORM-TRY-002 tryout inspection procedure (synthetic ref)",
            "Test Stage": "Virtual / CAE",
            "Build Phase": "Concept",
            "Sample Size": "3 tryout parts minimum",
            "Test Duration / Cycles": "Die tryout event",
            "Test Conditions": "Nominal and boundary material lot conditions",
        }
    if "modal" in text or "stiffness" in text:
        return {
            "Validation Level": "Subsystem",
            "Test Method / Procedure": "Modal analysis and local stiffness test",
            "Test Standard Reference": "NVH-MOD-005 modal test procedure (synthetic ref)",
            "Test Stage": "DV",
            "Build Phase": "Beta",
            "Sample Size": "1 body structure or subsystem",
            "Test Duration / Cycles": "Modal sweep and stiffness measurement",
            "Test Conditions": "Representative body boundary conditions",
        }
    return {
        "Validation Level": "Component",
        "Test Method / Procedure": "Engineering validation procedure to be defined",
        "Test Standard Reference": "TBD - assign approved internal test standard",
        "Test Stage": "DV",
        "Build Phase": "Alpha",
        "Sample Size": "TBD",
        "Test Duration / Cycles": "TBD",
        "Test Conditions": "Representative engineering condition",
    }


def dvp_row(
    dfmea_row_data: pd.Series,
    test: str,
    validation_type: str,
    objective: str,
    criteria: str,
    team: str,
    notes: str,
    validation_status: str = "Planned",
    pass_fail: str = "TBD",
    test_id: str = "",
) -> dict[str, Any]:
    severity = int(dfmea_row_data.get("Severity", 0) or 0)
    ap = str(dfmea_row_data.get("Action Priority", "Low"))
    confidence = 0.7
    source_type = "Synthetic MVP Rule"
    reason = human_review_reason(severity, ap, confidence, source_type)
    profile = validation_profile(test, validation_type)
    return {
        "Test ID": test_id,
        "Requirement ID": dfmea_row_data.get("Requirement ID", "REQ-GEN-001"),
        "Linked Failure Mode ID": dfmea_row_data.get("Failure Mode ID", ""),
        "Linked Failure Mode": dfmea_row_data.get("Potential Failure Mode", ""),
        "Recommended Validation Test": test,
        "Validation Type": validation_type,
        "Validation Level": profile["Validation Level"],
        "Test Objective": objective,
        "Test Method / Procedure": profile["Test Method / Procedure"],
        "Test Standard Reference": profile["Test Standard Reference"],
        "Test Stage": profile["Test Stage"],
        "Build Phase": profile["Build Phase"],
        "Sample Size": profile["Sample Size"],
        "Test Duration / Cycles": profile["Test Duration / Cycles"],
        "Test Conditions": profile["Test Conditions"],
        "Acceptance Criteria": criteria,
        "Planned Start Date": "TBD",
        "Planned Completion Date": "TBD",
        "Actual Completion Date": "",
        "Actual Result": "",
        "Pass / Fail": pass_fail,
        "Issue ID": "",
        "Evidence Link": "TBD",
        "Validation Status": validation_status,
        "Responsible Team": team,
        "Test Owner": team,
        "Notes": notes,
        "AI Suggestion ID": "",
        "AI Confidence Score": confidence,
        "Source Type": source_type,
        "Source Strength": "Low",
        "Citation Count": 0,
        "Human Review Required": review_required(reason),
        "Reason for Human Review": reason,
        "Future RAG Citation": "Not connected",
        "Source Evidence": "",
        "Source File": "",
        "Source Sheet": "",
        "Source Row": "",
        "Source Chunk ID": "",
        "AI Confidence": confidence,
        "Review Status": "Draft",
        "Engineer Decision": "Pending",
        "Rejection Reason": "",
        "Final Approved Text": f"{test}: {objective}",
        "Reviewed By": "TBD",
        "Review Date": "",
        "Approval Status": "Draft",
        "Approval Notes": "",
        "Change Log": "Initial validation recommendation generated by prototype.",
    }


def validation_rows_for_failure(row: pd.Series, settings: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    settings = settings or {}
    failure_mode = str(row.get("Potential Failure Mode", "")).strip()
    mode = failure_mode.lower()
    category = str(row.get("Risk Category", ""))
    severity = int(row.get("Severity", 0) or 0)
    rows: list[dict[str, Any]] = []

    def add(
        test: str,
        val_type: str,
        objective: str,
        criteria: str,
        team: str,
        notes: str,
        validation_status: str = "Planned",
        pass_fail: str = "TBD",
        test_id: str = "",
    ) -> None:
        rows.append(
            dvp_row(
                row,
                test,
                val_type,
                objective,
                criteria,
                team,
                notes,
                validation_status=validation_status,
                pass_fail=pass_fail,
                test_id=test_id,
            )
        )

    if "crash" in mode or "load-path" in mode or "load path" in mode or category == "Crash / Safety" or (
        severity >= 9 and any(term in mode for term in ["crash", "load path", "load-path", "deformation"])
    ):
        add(
            "Front crash load-path CAE correlation and physical crash validation",
            "CAE + Physical",
            "Confirm BIW reinforcement supports intended crash energy management",
            "Meets crash pulse, intrusion, deformation mode, and load-path targets",
            "BIW / Safety / CAE",
            "Required for crash or load-path risk. Crash risks are never intentionally left unlinked.",
        )

    if ("fatigue" in mode or "crack" in mode) and "spot weld" not in mode:
        add(
            "CAE durability analysis",
            "CAE",
            "Confirm fatigue margin under representative durability load cases",
            "Meets agreed durability margin and no high-risk stress concentration remains open",
            "BIW / CAE",
            "Use as early screening before physical validation.",
        )
        add(
            "Physical durability test with teardown inspection",
            "Physical",
            "Confirm component survives durability schedule without crack initiation",
            "No cracks, separated joints, or unacceptable permanent deformation",
            "BIW / Validation",
            "Teardown should focus on formed radii, beads, welds, and load-path transitions.",
        )

    if "spot weld fatigue" in mode:
        add(
            "Weld fatigue durability assessment",
            "CAE + Physical",
            "Confirm welded joint durability under representative load transfer conditions",
            "No weld fatigue, separation, or unacceptable joint degradation",
            "BIW / CAE / Validation",
            "Focused on cyclic load transfer rather than weld process conformance.",
        )
        add(
            "Weld teardown inspection after durability exposure",
            "Physical",
            "Confirm welded joints remain intact after durability loading",
            "No cracked welds, separated joints, or unacceptable heat affected zone damage",
            "BIW / Validation / Quality",
            "Use teardown evidence for high-load BIW joints.",
        )

    if "nugget" in mode or "quality" in mode or ("weld" in mode and "fatigue" not in mode):
        add(
            "Weld strength test and nugget inspection",
            "Physical / Manufacturing",
            "Confirm weld nugget size and joint strength meet engineering intent",
            "Nugget size, peel quality, and strength meet released criteria",
            "Manufacturing / Quality",
            "Include weld access and process window review.",
        )
        add(
            "Weld process audit",
            "Manufacturing",
            "Confirm weld process controls and access support consistent joint quality",
            "Process audit shows no unresolved access, shunting, or parameter concerns",
            "Manufacturing / Quality",
            "Connects weld quality risk to launch control plan.",
        )

    if "adhesive" in mode or "bond" in mode:
        add(
            "Adhesive lap shear and environmental aging validation",
            "Materials / Physical",
            "Confirm bonded joint strength after environmental exposure",
            "Bond strength and failure mode meet material and process targets",
            "Materials / Manufacturing",
            "Include surface prep and cure process assumptions.",
        )

    if "corrosion" in mode or "e-coat" in mode or "flange" in mode:
        add(
            "Corrosion cycle test",
            "Environmental",
            "Confirm long-term corrosion protection for exposed interfaces",
            "No unacceptable corrosion, perforation, or coating breakdown",
            "Materials / Corrosion",
            "Prioritize underbody, overlap, and water-trap areas.",
        )
        add(
            "E-coat coverage review with sectioning inspection",
            "Environmental / Manufacturing",
            "Confirm coating reaches closed sections and flange interfaces",
            "Coverage meets agreed minimum thickness and visual coverage criteria",
            "Materials / Corrosion / Manufacturing",
            "Use section cuts at representative restricted geometry.",
        )

    if "dimensional" in mode or "fit" in mode or "variation" in mode:
        add(
            "Body build dimensional study",
            "Manufacturing Validation",
            "Confirm assembly variation remains within tolerance intent",
            "Critical dimensions meet GD&T and build quality targets",
            "Dimensional Engineering",
            "Review datums, locators, fixture strategy, and measurement plan.",
        )

    if "nvh" in mode or "vibration" in mode or "stiffness" in mode:
        add(
            "Modal analysis and local stiffness validation",
            "CAE + Physical",
            "Confirm local modes and stiffness do not create objectionable vibration",
            "Modal and stiffness results meet agreed NVH targets",
            "NVH / BIW",
            "Correlate analysis with test where feasible.",
        )

    if "forming" in mode or "thinning" in mode or "wrinkling" in mode:
        add(
            "Forming simulation and tryout thickness inspection",
            "Manufacturing Validation",
            "Confirm part can be formed without unacceptable thinning or wrinkles",
            "Thickness, splits, wrinkles, and surface quality meet release targets",
            "Stamping / Manufacturing",
            "Connect findings to die changes or part geometry updates.",
        )

    if "pull-out" in mode or "fastener" in mode or "bracket" in mode:
        if settings.get("demo_gap_mode"):
            add(
                "Fastener pull-out and attachment durability validation",
                "CAE + Physical",
                "Confirm bracket and fastener attachment can withstand expected static and durability loads",
                "No pull-out, crack initiation, fastener loosening, or unacceptable permanent deformation",
                "BIW / Validation / Manufacturing",
                "AI-recommended closure for the open bracket or fastener pull-out validation gap.",
                validation_status="Proposed",
                pass_fail="TBD",
                test_id="TEST-014",
            )
            rows[-1]["Test Method / Procedure"] = (
                "Local CAE stress review, pull-out test, torque-to-failure test, "
                "post-durability teardown inspection"
            )
            rows[-1]["Build Phase"] = "Alpha / Beta"
            rows[-1]["Sample Size"] = "3 components minimum"
            rows[-1]["Test Duration / Cycles"] = "Based on applicable durability loading profile"
            rows[-1]["Test Conditions"] = (
                "Worst-case bracket load, tolerance stack-up, and representative attachment condition"
            )
            rows[-1]["Approval Status"] = "Under Review"
            rows[-1]["Engineer Decision"] = "Pending"
            rows[-1]["Reason for Human Review"] = "Proposed validation for open gap; engineer approval required"
            rows[-1]["Human Review Required"] = "Yes"
            return rows
        add(
            "Fastener pull-out and attachment durability test",
            "Physical",
            "Confirm attachment strength and durability margin",
            "No pull-out, crack initiation, or unacceptable permanent deformation",
            "BIW / Validation",
            "Include worst-case load path and assembly variation assumptions.",
        )
        add(
            "Local CAE stiffness and stress review",
            "CAE",
            "Confirm local bracket and fastener load spread has acceptable margin",
            "CAE stress and stiffness results meet engineering target",
            "BIW / CAE",
            "Complements physical attachment testing.",
        )

    if "service access" in mode:
        add(
            "Service access mockup review",
            "Physical / Review",
            "Confirm service tool access and removal sequence are feasible",
            "Service operation can be completed with approved tool clearance",
            "BIW / Service / Manufacturing",
            "Supports serviceability risk closure.",
        )

    if not rows and severity >= 8:
        add(
            "Engineering validation plan review",
            "Review",
            "Define validation coverage for high-severity risk",
            "Responsible team approves coverage plan",
            "BIW / Validation",
            "High-severity risks need explicit validation linkage.",
        )

    return rows


def generate_dvp(dfmea_df: pd.DataFrame, settings: dict[str, Any] | None = None) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if dfmea_df.empty:
        return empty_df(DVP_COLUMNS)

    for _, row in dfmea_df.iterrows():
        rows.extend(validation_rows_for_failure(row, settings))

    used_ids = {str(row.get("Test ID", "")).strip() for row in rows if str(row.get("Test ID", "")).strip()}
    next_id = 1
    for index, row in enumerate(rows, start=1):
        if not str(row.get("Test ID", "")).strip():
            while f"TEST-{next_id:03d}" in used_ids:
                next_id += 1
            row["Test ID"] = f"TEST-{next_id:03d}"
            used_ids.add(row["Test ID"])
            next_id += 1
        row["AI Suggestion ID"] = f"AI-DVPR-{index:03d}"

    return sort_dvp_by_test_id(pd.DataFrame(rows, columns=DVP_COLUMNS))


def test_id_sort_value(test_id: Any) -> int:
    value = str(test_id or "").strip().upper()
    if value.startswith("TEST-"):
        value = value.replace("TEST-", "", 1)
    return int(value) if value.isdigit() else 999999


def sort_dvp_by_test_id(dvp_df: pd.DataFrame) -> pd.DataFrame:
    if dvp_df.empty or "Test ID" not in dvp_df:
        return dvp_df
    sorted_df = dvp_df.copy()
    sorted_df["_test_sort"] = sorted_df["Test ID"].map(test_id_sort_value)
    sorted_df = sorted_df.sort_values(["_test_sort", "Test ID"], kind="stable").drop(columns=["_test_sort"])
    return sorted_df.reset_index(drop=True)


def coverage_assessment(dfmea_row_data: pd.Series, matches: pd.DataFrame) -> dict[str, Any]:
    severity = int(dfmea_row_data.get("Severity", 0) or 0)
    mode = str(dfmea_row_data.get("Potential Failure Mode", "")).lower()
    category = str(dfmea_row_data.get("Risk Category", ""))

    if matches.empty:
        return {
            "Coverage Status": "Gap",
            "Coverage Score": 0,
            "Coverage Reason": "No linked DVP&R validation item exists.",
            "Missing Validation Type": "CAE + Physical" if severity >= 8 else "Validation method",
            "Recommended Additional Test": "Add a linked validation test and assign accountable owner.",
            "Gap Description": "Failure mode has no linked validation item.",
        }

    val_text = " ".join(
        str(value).lower()
        for value in matches[["Validation Type", "Validation Level", "Recommended Validation Test"]].fillna("").to_numpy().flatten()
    )
    has_cae = "cae" in val_text
    has_physical = any(term in val_text for term in ["physical", "teardown", "pull-out", "crash validation"])
    has_acceptance = matches["Acceptance Criteria"].astype(str).str.strip().ne("").all()
    has_evidence = matches["Evidence Link"].astype(str).str.strip().ne("").all()
    proposed = matches["Validation Status"].astype(str).str.contains("Proposed", case=False, na=False).any()

    if proposed:
        proposed_test = str(matches.iloc[0].get("Recommended Validation Test", "AI-recommended validation test"))
        return {
            "Coverage Status": "Proposed Coverage",
            "Coverage Score": 70,
            "Coverage Reason": "AI-recommended validation exists, but engineer approval and test evidence are still required.",
            "Missing Validation Type": "Engineer approval and execution evidence",
            "Recommended Additional Test": proposed_test,
            "Gap Description": "Before AI review this risk had no linked validation. AI has proposed a closure test.",
        }

    if (category == "Crash / Safety" or "crash" in mode or "load-path" in mode or "load path" in mode) and not (
        has_cae and has_physical
    ):
        return {
            "Coverage Status": "Partial",
            "Coverage Score": 50,
            "Coverage Reason": "Crash risk requires linked CAE and physical validation evidence.",
            "Missing Validation Type": "CAE + Physical crash validation",
            "Recommended Additional Test": "Add crash CAE correlation and physical crash validation.",
            "Gap Description": "Crash validation coverage is incomplete.",
        }

    if severity >= 8 and not (has_cae and has_physical):
        return {
            "Coverage Status": "Partial",
            "Coverage Score": 50,
            "Coverage Reason": "High-severity risk is not covered by both CAE and physical validation.",
            "Missing Validation Type": "Independent CAE or physical confirmation",
            "Recommended Additional Test": "Add independent CAE, physical, or teardown confirmation.",
            "Gap Description": "High-severity validation coverage is partial.",
        }

    if not has_acceptance:
        return {
            "Coverage Status": "Partial",
            "Coverage Score": 50,
            "Coverage Reason": "Linked validation exists but acceptance criteria are missing.",
            "Missing Validation Type": "Acceptance criteria",
            "Recommended Additional Test": "Define measurable acceptance criteria.",
            "Gap Description": "Validation acceptance criteria are missing.",
        }

    if not has_evidence:
        score = 100 if has_cae and has_physical else 70
        return {
            "Coverage Status": "Covered",
            "Coverage Score": score,
            "Coverage Reason": "Validation is planned. Evidence link remains TBD until execution.",
            "Missing Validation Type": "",
            "Recommended Additional Test": "Attach evidence after test execution.",
            "Gap Description": "",
        }

    return {
        "Coverage Status": "Covered",
        "Coverage Score": 100 if has_cae and has_physical else 70,
        "Coverage Reason": "Linked validation coverage exists.",
        "Missing Validation Type": "",
        "Recommended Additional Test": "",
        "Gap Description": "",
    }


def generate_traceability(dfmea_df: pd.DataFrame, dvp_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if dfmea_df.empty:
        return empty_df(TRACE_COLUMNS)

    for _, dfmea_row_data in dfmea_df.iterrows():
        fm_id = str(dfmea_row_data.get("Failure Mode ID", "")).strip()
        matches = pd.DataFrame()
        if not dvp_df.empty and fm_id:
            matches = dvp_df[dvp_df["Linked Failure Mode ID"].astype(str).str.strip() == fm_id]

        tests = [str(value).strip() for value in matches.get("Recommended Validation Test", pd.Series(dtype=str)).tolist()]
        test_ids = [str(value).strip() for value in matches.get("Test ID", pd.Series(dtype=str)).tolist()]
        assessment = coverage_assessment(dfmea_row_data, matches)
        status = assessment["Coverage Status"]
        proposed_matches = matches[
            matches.get("Validation Status", pd.Series(dtype=str)).astype(str).str.contains("Proposed", case=False, na=False)
        ] if not matches.empty else pd.DataFrame()
        ai_test_id = ""
        ai_test_name = ""
        if not proposed_matches.empty:
            ai_test_id = str(proposed_matches.iloc[0].get("Test ID", ""))
            ai_test_name = str(proposed_matches.iloc[0].get("Recommended Validation Test", ""))
        elif status in {"Gap", "Partial"}:
            ai_test_name = assessment["Recommended Additional Test"]
        expected_improvement = "0% to 70%" if status == "Proposed Coverage" else ("0% to 70% after AI recommendation" if status == "Gap" else "")
        severity = int(dfmea_row_data.get("Severity", 0) or 0)
        ap = str(dfmea_row_data.get("Action Priority", "Low"))
        confidence = 0.7
        source_type = "Synthetic MVP Rule"
        reason = human_review_reason(severity, ap, confidence, source_type, status)

        rows.append(
            {
                "Requirement ID": dfmea_row_data.get("Requirement ID", ""),
                "Function ID": dfmea_row_data.get("Function ID", ""),
                "Failure Mode ID": fm_id,
                "Failure Mode": dfmea_row_data.get("Potential Failure Mode", ""),
                "Risk Category": dfmea_row_data.get("Risk Category", ""),
                "Severity": severity,
                "RPN": int(dfmea_row_data.get("RPN", 0) or 0),
                "Action Priority": ap,
                "Linked Test IDs": "; ".join(test_ids) if test_ids else "Not linked",
                "Linked DVP&R Test": "; ".join(tests) if tests else "Not linked",
                "Coverage Status": status,
                "Coverage Score": assessment["Coverage Score"],
                "Coverage Reason": assessment["Coverage Reason"],
                "Missing Validation Type": assessment["Missing Validation Type"],
                "Recommended Additional Test": assessment["Recommended Additional Test"],
                "AI Recommended Test ID": ai_test_id,
                "AI Recommended Test Name": ai_test_name,
                "Before AI Review Coverage Status": "Gap" if status == "Proposed Coverage" else status,
                "After AI Recommendation Coverage Status": status,
                "Expected Coverage Improvement": expected_improvement,
                "Gap Description": assessment["Gap Description"],
                "Recommended Action": dfmea_row_data.get("Recommended Action", ""),
                "Priority": "High" if status != "Covered" and severity >= 8 else ("Medium" if status != "Covered" else "Normal"),
                "Owner": dfmea_row_data.get("Responsible Team", "BIW / Validation"),
                "AI Suggestion ID": f"AI-TRACE-{len(rows) + 1:03d}",
                "AI Confidence Score": confidence,
                "Source Type": source_type,
                "Source Strength": "Low",
                "Citation Count": 0,
                "Human Review Required": review_required(reason),
                "Reason for Human Review": reason,
                "Future RAG Citation": "Not connected",
                "Engineer Decision": "Pending",
                "Rejection Reason": "",
                "Final Approved Text": assessment["Coverage Reason"],
                "Reviewed By": "TBD",
                "Review Date": "",
                "Approval Status": "Draft",
                "Approval Notes": "",
                "Change Log": "Traceability generated from linked failure mode and validation test IDs.",
            }
        )

    return pd.DataFrame(rows, columns=TRACE_COLUMNS)


def generate_gap_analysis(trace_df: pd.DataFrame) -> pd.DataFrame:
    if trace_df.empty:
        return empty_df(GAP_COLUMNS)

    rows: list[dict[str, Any]] = []
    for _, row in trace_df.iterrows():
        status = str(row.get("Coverage Status", ""))
        if status == "Covered":
            continue
        if status == "Gap":
            gap_type = "Missing Validation Test"
        elif status == "Proposed Coverage":
            gap_type = "Proposed Validation Closure"
        else:
            gap_type = "Partial Validation Coverage"
        closure_status = "Proposed" if status == "Proposed Coverage" else "Open"
        final_text = row.get("AI Recommended Test Name", "") or row.get("Recommended Additional Test", "")
        evidence_link = ""
        final_gap_status = closure_status
        rows.append(
            {
                "Gap ID": f"GAP-{len(rows) + 1:03d}",
                "Failure Mode ID": row.get("Failure Mode ID", ""),
                "Failure Mode": row.get("Failure Mode", ""),
                "Severity": row.get("Severity", ""),
                "RPN": row.get("RPN", ""),
                "Action Priority": row.get("Action Priority", ""),
                "Risk Category": row.get("Risk Category", ""),
                "Gap Type": gap_type,
                "Gap Description": row.get("Gap Description", "") or row.get("Coverage Reason", ""),
                "Recommended Fix": row.get("Recommended Additional Test", "") or row.get("Recommended Action", ""),
                "Responsible Team": row.get("Owner", "BIW / Validation"),
                "Priority": row.get("Priority", "Medium"),
                "Status": closure_status,
                "Gap Status": closure_status,
                "AI Recommended Test ID": row.get("AI Recommended Test ID", ""),
                "AI Recommended Test Name": row.get("AI Recommended Test Name", ""),
                "AI Recommended Closure Test": final_text,
                "Expected Coverage Improvement": row.get("Expected Coverage Improvement", ""),
                "After Engineer Approval Coverage Status": "Covered if accepted and evidence is attached",
                "Engineer Decision": "Pending",
                "Gap Closure Status": closure_status,
                "Evidence Link": evidence_link,
                "Final Gap Closure Status": final_gap_status,
                "Rejection Reason": "Not applicable",
                "Final Approved Text": final_text,
                "Reviewed By": "TBD",
                "Review Date": "",
                "Approval Status": "Under Review" if status == "Proposed Coverage" else "Draft",
                "Change Log": "Gap row generated from traceability coverage status.",
            }
        )
    return pd.DataFrame(rows, columns=GAP_COLUMNS)


LESSON_LIBRARY = [
    {
        "keywords": ["formed radius", "crack at formed radius"],
        "Prior Issue Type": "Durability Failure",
        "Lesson Learned": "Small radii and abrupt section changes can increase local strain and fatigue risk.",
        "Recommended Design Action": "Review local radius, gauge, bead strategy, and CAE stress concentration.",
        "Recommended Validation Action": "Run CAE durability and physical durability teardown inspection.",
    },
    {
        "keywords": ["spot weld fatigue", "spot weld quality", "nugget"],
        "Prior Issue Type": "Weld Quality Issue",
        "Lesson Learned": "Poor weld pitch, access, or nugget consistency can reduce joint durability.",
        "Recommended Design Action": "Review weld pitch, access, nugget size, flange width, and load distribution.",
        "Recommended Validation Action": "Add weld fatigue assessment, nugget inspection, and teardown review.",
    },
    {
        "keywords": ["adhesive", "bond"],
        "Prior Issue Type": "Manufacturing Launch Issue",
        "Lesson Learned": "Surface prep, bond gap, and cure process variation can affect bonded joint performance.",
        "Recommended Design Action": "Review adhesive bead path, gap control, surface prep, and process window.",
        "Recommended Validation Action": "Run lap shear and environmental aging validation.",
    },
    {
        "keywords": ["corrosion", "flange", "e-coat"],
        "Prior Issue Type": "Corrosion Issue",
        "Lesson Learned": "Poor drainage and restricted coating access can create long-term corrosion risk.",
        "Recommended Design Action": "Review drainage, sealer path, e-coat access, and water traps.",
        "Recommended Validation Action": "Run corrosion cycle and e-coat coverage sectioning.",
    },
    {
        "keywords": ["dimensional", "variation", "fit"],
        "Prior Issue Type": "Dimensional Build Issue",
        "Lesson Learned": "Weak datum strategy and tolerance stack sensitivity can drive build variation.",
        "Recommended Design Action": "Review GD&T, locating strategy, fixture variation, and body framing sequence.",
        "Recommended Validation Action": "Run body build dimensional study.",
    },
    {
        "keywords": ["crash", "load-path", "load path", "instability"],
        "Prior Issue Type": "Crash Performance Issue",
        "Lesson Learned": "Local section instability or weak joint continuity can change crash energy management.",
        "Recommended Design Action": "Review load-path continuity, section stability, and CAE-to-test correlation.",
        "Recommended Validation Action": "Run crash CAE correlation and physical crash validation.",
    },
    {
        "keywords": ["nvh", "vibration", "stiffness"],
        "Prior Issue Type": "NVH Issue",
        "Lesson Learned": "Unsupported spans and compliant joints can create noise or vibration issues.",
        "Recommended Design Action": "Review local stiffness, attachment strategy, and modal response.",
        "Recommended Validation Action": "Run modal analysis and local stiffness validation.",
    },
    {
        "keywords": ["forming", "thinning", "wrinkling"],
        "Prior Issue Type": "Stamping/Formability Issue",
        "Lesson Learned": "Aggressive draw geometry can create thinning, wrinkles, splits, or scrap risk.",
        "Recommended Design Action": "Review draw strategy, radius, bead placement, material flow, and trim line.",
        "Recommended Validation Action": "Run forming simulation and tryout thickness inspection.",
    },
]


def generate_lessons(dfmea_df: pd.DataFrame) -> pd.DataFrame:
    if dfmea_df.empty:
        return empty_df(LESSON_COLUMNS)

    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for _, row in dfmea_df.iterrows():
        failure_mode = str(row.get("Potential Failure Mode", "")).strip()
        mode_text = failure_mode.lower()
        for lesson in LESSON_LIBRARY:
            if any(keyword in mode_text for keyword in lesson["keywords"]):
                key = (str(row.get("Failure Mode ID", "")), lesson["Prior Issue Type"])
                if key in seen:
                    continue
                seen.add(key)
                severity = int(row.get("Severity", 0) or 0)
                ap = str(row.get("Action Priority", "Low"))
                confidence = 0.7
                source_type = "Synthetic MVP Rule"
                reason = human_review_reason(severity, ap, confidence, source_type)
                rows.append(
                    {
                        "Lesson ID": f"LL-{len(rows) + 1:03d}",
                        "Related Failure Mode ID": row.get("Failure Mode ID", ""),
                        "Related Requirement ID": row.get("Requirement ID", ""),
                        "Risk Category": row.get("Risk Category", ""),
                        "Related Failure Mode": failure_mode,
                        "Lesson Learned": lesson["Lesson Learned"],
                        "Prior Issue Type": lesson["Prior Issue Type"],
                        "Applicability to Current Component": "Applicable based on generated risk category and component inputs.",
                        "Recommended Design Action": lesson["Recommended Design Action"],
                        "Recommended Validation Action": lesson["Recommended Validation Action"],
                        "Source Type": source_type,
                        "Source Confidence": "Medium",
                        "Engineer Decision": "Pending",
                        "AI Suggestion ID": f"AI-LESSON-{len(rows) + 1:03d}",
                        "AI Confidence Score": confidence,
                        "Source Strength": "Low",
                        "Citation Count": 0,
                        "Human Review Required": review_required(reason),
                        "Reason for Human Review": reason,
                        "Future RAG Citation": "Not connected",
                        "Source Evidence": "",
                        "Source File": "",
                        "Source Sheet": "",
                        "Source Row": "",
                        "Source Chunk ID": "",
                        "AI Confidence": confidence,
                        "Review Status": "Draft",
                        "Rejection Reason": "",
                        "Final Approved Text": lesson["Lesson Learned"],
                        "Reviewed By": "TBD",
                        "Review Date": "",
                        "Approval Status": "Draft",
                        "Approval Notes": "",
                        "Change Log": "Lesson generated from synthetic lessons library.",
                    }
                )
    return pd.DataFrame(rows, columns=LESSON_COLUMNS)


def generate_all() -> None:
    inputs = component_inputs()
    settings = generation_settings()
    dfmea_df = generate_dfmea(inputs)
    dvp_df = generate_dvp(dfmea_df, settings)
    trace_df = generate_traceability(dfmea_df, dvp_df)
    gap_df = generate_gap_analysis(trace_df)
    lessons_df = generate_lessons(dfmea_df)
    dfmea_df, dvp_df, lessons_df, retrieved_df = ground_with_rag(dfmea_df, dvp_df, lessons_df, inputs)
    if st.session_state.get("llm_enrich_mode") and not RAG_IMPORT_ERROR:
        try:
            dfmea_df, dvp_df, llm_notes = apply_llm_enrichment(dfmea_df, dvp_df, inputs)
            st.session_state.llm_notes = llm_notes
        except Exception as exc:
            st.session_state.llm_notes = [f"LLM enrichment skipped: {exc}"]
    trace_df = generate_traceability(dfmea_df, dvp_df)
    gap_df = generate_gap_analysis(trace_df)
    gap_df = augment_gap_analysis(gap_df, dfmea_df, dvp_df, retrieved_df)
    st.session_state.retrieved_sources_df = retrieved_df
    st.session_state.pdiag_df = generate_p_diagram(inputs)
    st.session_state.dfmea_df = dfmea_df
    st.session_state.dvp_df = dvp_df
    st.session_state.trace_df = trace_df
    st.session_state.gap_df = gap_df
    st.session_state.lessons_df = lessons_df
    st.session_state.last_generated = True


def refresh_downstream_from_edits() -> None:
    st.session_state.trace_df = generate_traceability(st.session_state.dfmea_df, st.session_state.dvp_df)
    gap_df = generate_gap_analysis(st.session_state.trace_df)
    st.session_state.gap_df = augment_gap_analysis(
        gap_df,
        st.session_state.dfmea_df,
        st.session_state.dvp_df,
        st.session_state.get("retrieved_sources_df", empty_df(RETRIEVED_SOURCES_COLUMNS)),
    )
    st.session_state.lessons_df = generate_lessons(st.session_state.dfmea_df)


def dataframe_to_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def safe_count(df: pd.DataFrame, column: str, value: str) -> int:
    if df.empty or column not in df:
        return 0
    return int((df[column] == value).sum())


def mean_numeric(df: pd.DataFrame, column: str) -> float:
    if df.empty or column not in df:
        return 0.0
    return float(pd.to_numeric(df[column], errors="coerce").fillna(0).mean())


def dashboard_dataframe(
    dfmea_df: pd.DataFrame,
    dvp_df: pd.DataFrame,
    trace_df: pd.DataFrame,
    gap_df: pd.DataFrame,
    lessons_df: pd.DataFrame,
) -> pd.DataFrame:
    high_dfmea = dfmea_df[dfmea_df["Severity"] >= 8] if not dfmea_df.empty else dfmea_df
    coverage = mean_numeric(trace_df, "Coverage Score")
    high_coverage = mean_numeric(trace_df[trace_df["Severity"] >= 8], "Coverage Score") if not trace_df.empty else 0
    avg_initial = mean_numeric(dfmea_df, "Initial RPN")
    avg_revised = mean_numeric(dfmea_df, "Revised RPN")
    reduction = round((avg_initial - avg_revised) / avg_initial, 3) if avg_initial else 0
    accepted = safe_count(dfmea_df, "Engineer Decision", "Accept") + safe_count(dvp_df, "Engineer Decision", "Accept")
    rejected = safe_count(dfmea_df, "Engineer Decision", "Reject") + safe_count(dvp_df, "Engineer Decision", "Reject")

    prototype_rows = [
        ("Prototype Boundary", "Prototype Mode", "Synthetic Demo Data", ""),
        ("Prototype Boundary", "Data Source", "Generic BIW Rules / No Confidential Data", ""),
        ("Prototype Boundary", "Production Intent", "Secure Internal RAG", ""),
        ("Prototype Boundary", "Engineer Approval Required", "Yes", ""),
        ("Prototype Boundary", "AI Role", "Draft / Recommend / Check Gaps / Summarize", ""),
        ("Prototype Boundary", "Prototype Boundary Statement", PROTOTYPE_BOUNDARY_STATEMENT, ""),
        ("Prototype Boundary", "Gap Closure Rule", GAP_WORKFLOW_NOTE, ""),
    ]

    kpis = [
        ("KPI", "Total DFMEA Failure Modes", len(dfmea_df), ""),
        ("KPI", "High-Severity Failure Modes", len(high_dfmea), ""),
        ("KPI", "Highest RPN", int(dfmea_df["RPN"].max()) if not dfmea_df.empty else 0, ""),
        ("KPI", "Average Initial RPN", round(avg_initial, 1), ""),
        ("KPI", "Average Revised RPN", round(avg_revised, 1), ""),
        ("KPI", "Estimated RPN Reduction %", reduction, ""),
        ("KPI", "High Action Priority Items", safe_count(dfmea_df, "Action Priority", "High"), ""),
        ("KPI", "Total DVP&R Tests", len(dvp_df), ""),
        ("KPI", "Overall Validation Coverage Score", round(coverage / 100, 3), ""),
        ("KPI", "High-Severity Validation Coverage Score", round(high_coverage / 100, 3), ""),
        ("KPI", "Open High-Priority Gaps", len(gap_df[gap_df["Priority"].eq("High")]) if not gap_df.empty else 0, ""),
        ("KPI", "Open Validation Gaps", len(gap_df), ""),
        ("KPI", "Partial Coverage Items", safe_count(trace_df, "Coverage Status", "Partial"), ""),
        ("KPI", "Open Recommended Actions", safe_count(dfmea_df, "Action Status", "Open"), ""),
        ("KPI", "Lessons Reused", len(lessons_df), ""),
        ("KPI", "AI Suggestions Accepted", accepted, ""),
        ("KPI", "AI Suggestions Rejected", rejected, ""),
    ]

    footnote_rows = [
        ("Dashboard Footnote", f"Footnote {idx}", note, "")
        for idx, note in enumerate(DASHBOARD_FOOTNOTES, start=1)
    ]

    coverage_score_rows = [
        ("Coverage Score Header", "Coverage Measure", "Validation Coverage Score", ""),
        ("Coverage Score", "Overall Validation Coverage Score", round(coverage / 100, 3), ""),
        ("Coverage Score", "High-Severity Coverage Score", round(high_coverage / 100, 3), ""),
    ]

    rpn_rows = [("RPN Chart Header", "Failure Mode", "Initial RPN", "Revised RPN")]
    if not dfmea_df.empty:
        for _, row in dfmea_df.sort_values("Initial RPN", ascending=False).iterrows():
            rpn_rows.append(
                (
                    "RPN Comparison",
                    f"{row['Failure Mode ID']} {row['Potential Failure Mode']}",
                    int(row["Initial RPN"]),
                    int(row["Revised RPN"]),
                )
            )

    category_rows = [("Risk Category Header", "Risk Category", "Failure Mode Count", "")]
    if not dfmea_df.empty:
        category_counts = dfmea_df["Risk Category"].value_counts().sort_index()
        category_rows = [("Risk Category Count", category, int(count), "") for category, count in category_counts.items()]
        category_rows.insert(0, ("Risk Category Header", "Risk Category", "Failure Mode Count", ""))

    coverage_rows = [("Coverage Status Header", "Coverage Status", "Failure Mode Count", "")]
    if not trace_df.empty:
        coverage_counts = trace_df["Coverage Status"].value_counts().sort_index()
        coverage_rows = [("Coverage Status Count", status, int(count), "") for status, count in coverage_counts.items()]
        coverage_rows.insert(0, ("Coverage Status Header", "Coverage Status", "Failure Mode Count", ""))

    owner_rows = [("Open Actions Header", "Action Owner", "Open Action Count", "")]
    if not dfmea_df.empty:
        owner_counts = dfmea_df[dfmea_df["Action Status"].eq("Open")]["Action Owner"].value_counts().sort_index()
        owner_rows = [("Open Actions by Owner", owner, int(count), "") for owner, count in owner_counts.items()]
        owner_rows.insert(0, ("Open Actions Header", "Action Owner", "Open Action Count", ""))

    # RAG knowledge-layer KPIs (roadmap Phase 13) - exported with the Dashboard sheet
    rag_rows: list[tuple[str, str, Any, str]] = []
    if not RAG_IMPORT_ERROR:
        try:
            stats = get_rag_store().get_collection_stats()
            retrieved_df = st.session_state.get("retrieved_sources_df", empty_df(RETRIEVED_SOURCES_COLUMNS))
            dfmea_grounded = int((dfmea_df.get("Source Chunk ID", pd.Series(dtype=str)).astype(str) != "").sum()) if not dfmea_df.empty else 0
            dvp_grounded = int((dvp_df.get("Source Chunk ID", pd.Series(dtype=str)).astype(str) != "").sum()) if not dvp_df.empty else 0
            total_rows = len(dfmea_df) + len(dvp_df)
            fallback_rows = total_rows - dfmea_grounded - dvp_grounded
            high_risk_no_source = (
                int(((dfmea_df["Severity"] >= 8) & (dfmea_df.get("Source Chunk ID", "").astype(str) == "")).sum())
                if not dfmea_df.empty and "Source Chunk ID" in dfmea_df.columns
                else 0
            )
            avg_conf = 0.0
            if total_rows:
                avg_conf = round(
                    (mean_numeric(dfmea_df, "AI Confidence") * len(dfmea_df) + mean_numeric(dvp_df, "AI Confidence") * len(dvp_df))
                    / total_rows,
                    2,
                )
            sources_used = int((retrieved_df["Used In Output"].astype(str) == "Yes").sum()) if not retrieved_df.empty else 0
            rag_rows = [
                ("RAG KPI", "Total Documents Indexed", stats["documents"], ""),
                ("RAG KPI", "Total Knowledge Chunks", stats["chunks"], ""),
                ("RAG KPI", "RAG Sources Used", sources_used, ""),
                ("RAG KPI", "DFMEA Rows with Source Evidence", dfmea_grounded, ""),
                ("RAG KPI", "DVP&R Rows with Source Evidence", dvp_grounded, ""),
                ("RAG KPI", "Rows Using Rule-Based Fallback", fallback_rows, ""),
                ("RAG KPI", "High-Risk Items Without Source Evidence", high_risk_no_source, ""),
                ("RAG KPI", "Average AI Confidence", avg_conf, ""),
            ]
        except Exception:
            rag_rows = []

    return pd.DataFrame(
        prototype_rows + kpis + rag_rows + footnote_rows + coverage_score_rows + rpn_rows + category_rows + coverage_rows + owner_rows,
        columns=["Section", "Metric", "Value", "Comparison Value"],
    )


def component_input_dataframe(inputs: dict[str, str]) -> pd.DataFrame:
    rows = [
        ("Component Name", inputs.get("component_name", "")),
        ("Vehicle Area", inputs.get("vehicle_area", "")),
        ("Component Category", inputs.get("component_category", "")),
        ("Material", inputs.get("material", "")),
        ("Thickness", inputs.get("thickness", "")),
        ("Joining Method", inputs.get("joining_method", "")),
        ("Manufacturing Process", inputs.get("manufacturing_process", "")),
        ("Function", inputs.get("primary_function", "")),
        ("Interfaces", inputs.get("interfaces", "")),
        ("Load Cases", inputs.get("load_cases", "")),
        ("Environmental Exposure", inputs.get("environmental_exposure", "")),
        ("Known Concerns", inputs.get("known_design_concerns", "")),
        ("Program Phase", inputs.get("program_phase", "")),
        ("Engineer Name", inputs.get("engineer_name", "")),
        ("Date Generated", date.today().isoformat()),
        ("Tool Version", TOOL_VERSION),
        ("Disclaimer", DISCLAIMER),
    ]
    return pd.DataFrame(rows, columns=["Field", "Value"])


def management_summary_dataframe(
    inputs: dict[str, str],
    dfmea_df: pd.DataFrame,
    dvp_df: pd.DataFrame,
    trace_df: pd.DataFrame,
    gap_df: pd.DataFrame,
    lessons_df: pd.DataFrame,
) -> pd.DataFrame:
    avg_initial = mean_numeric(dfmea_df, "Initial RPN")
    avg_revised = mean_numeric(dfmea_df, "Revised RPN")
    reduction = round((avg_initial - avg_revised) / avg_initial, 3) if avg_initial else 0
    high_severity = len(dfmea_df[dfmea_df["Severity"] >= 8]) if not dfmea_df.empty else 0
    high_priority_gaps = len(gap_df[gap_df["Priority"].eq("High")]) if not gap_df.empty else 0
    rows = [
        (
            "Executive Summary",
            "This proof of concept demonstrates an AI-assisted workflow for BIW sheet metal DFMEA and DVP&R development.",
        ),
        (
            "Tool Workflow",
            "The tool takes structured component inputs and generates draft DFMEA failure modes, Severity, Occurrence, Detection, RPN, Action Priority, recommended actions, revised RPN, DVP&R validation recommendations, requirement-to-risk-to-test traceability, gap analysis, lessons learned suggestions, pilot metrics, and a management dashboard.",
        ),
        (
            "Prototype Demonstration",
            "The goal is not to replace engineering judgment. The goal is to reduce repetitive documentation effort, improve validation traceability, preserve lessons learned, and help engineers identify gaps earlier.",
        ),
        (
            "Engineer Approval Boundary",
            "The engineer remains the final approver.",
        ),
        (
            "Demo Story",
            DEMO_STORY,
        ),
        (
            "Demo Story Takeaway",
            "This shows that the AI assistant does more than generate text. It supports risk prioritization, validation planning, traceability, and engineering review.",
        ),
        ("Key Finding", f"{len(dfmea_df)} DFMEA failure modes generated."),
        ("Key Finding", f"{high_severity} high-severity risks identified."),
        ("Key Finding", f"{len(dvp_df)} DVP&R validation items recommended or proposed."),
        ("Key Finding", f"Estimated RPN reduction of {reduction:.1%}."),
        ("Key Finding", f"{high_priority_gaps} high-priority validation gap or proposed closure item identified."),
        ("Key Finding", f"{len(lessons_df)} lessons learned reused."),
        (
            "RAG Knowledge Layer",
            "This prototype includes a local RAG knowledge layer. Engineering documents are parsed, "
            "chunked, embedded, and searched before recommendations are finalized. Rows with retrieved "
            "source evidence carry file, sheet, row, chunk ID, source strength, and AI confidence. Rows "
            "without matching sources are labeled as rule-based fallback and require engineer review.",
        ),
        (
            "Gap Closure Workflow",
            GAP_WORKFLOW_NOTE,
        ),
        (
            "Prototype Boundary",
            PROTOTYPE_BOUNDARY_STATEMENT,
        ),
        (
            "Action Priority Boundary",
            ACTION_PRIORITY_DISCLAIMER,
        ),
        (
            "Engineering Control",
            "The AI assistant only drafts, recommends, links, and flags gaps. The responsible engineer remains the final approver for all DFMEA and DVP&R content.",
        ),
        (
            "Final Pitch Positioning",
            FINAL_PITCH_POSITIONING,
        ),
        (
            "Next Step",
            "Approve a limited 8-12 week pilot using anonymized or approved internal BIW data and compare manual vs AI-assisted DFMEA/DVP&R preparation time and quality.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Section", "Summary"])


def settings_dataframe() -> pd.DataFrame:
    rows = [
        ("Rating Scale", "Severity", "1 low impact to 10 safety or regulatory impact"),
        ("Rating Scale", "Occurrence", "1 unlikely to 10 frequent"),
        ("Rating Scale", "Detection", "1 likely to detect to 10 unlikely to detect"),
        ("Formula", "RPN", "Severity x Occurrence x Detection"),
        ("Formula", "Initial RPN", "Initial Severity x Initial Occurrence x Initial Detection"),
        ("Formula", "Revised RPN", "Revised Severity x Revised Occurrence x Revised Detection"),
        ("Formula", "RPN Reduction %", "(Initial RPN - Revised RPN) / Initial RPN"),
        (
            "Action Priority Logic",
            "High",
            "Severity >= 9, or Severity >= 8 and Occurrence >= 4, or Severity >= 7 and RPN >= 120",
        ),
        ("Action Priority Logic", "Medium", "RPN >= 80 and not High"),
        ("Action Priority Logic", "Low", "All remaining rows"),
        ("AP (AIAG-VDA) Logic", "H / M / L", "Banded implementation of the AIAG-VDA 2019 Action Priority table driven by Severity, then Occurrence, then Detection - not an RPN threshold. Production must adopt the approved company AP table."),
        ("Special Characteristic Logic", "YC (Potential Critical)", "Severity 9-10: potential safety or regulatory effect - candidate critical characteristic"),
        ("Special Characteristic Logic", "SC (Potential Significant)", "Severity 5-8 with Occurrence >= 4 - candidate significant characteristic"),
        ("Special Characteristic Logic", "Note", "Final YC/CC/SC designation belongs to the responsible engineer and the program special characteristics process."),
        ("Test Stage Values", "Values", ", ".join(v for v in TEST_STAGE_VALUES if v)),
        ("Coverage Score Logic", "100", "Covered with CAE and physical validation"),
        ("Coverage Score Logic", "70", "Covered with one strong validation method"),
        ("Coverage Score Logic", "50", "Partial coverage"),
        ("Coverage Score Logic", "0", "No linked validation"),
        ("Coverage Score Logic", "Note", "Coverage score reflects validation maturity: covered, proposed, partial, or missing validation coverage."),
        ("Gap Status Logic", "Open", "No validation test is linked"),
        ("Gap Status Logic", "Proposed", "AI has proposed a closure test, but engineer has not accepted it"),
        ("Gap Status Logic", "Accepted", "Engineer accepted the proposed validation test"),
        ("Gap Status Logic", "In Progress", "Validation test is planned or being executed"),
        ("Gap Status Logic", "Closed", "Validation evidence is linked and accepted"),
        ("Gap Status Logic", "Waived", "Engineer-approved waiver with rationale"),
        ("Prototype Boundary", "Prototype Statement", PROTOTYPE_BOUNDARY_STATEMENT),
        ("Action Priority Boundary", "Prototype Disclaimer", ACTION_PRIORITY_DISCLAIMER),
        ("Gap Closure Boundary", "Engineer Acceptance Required", GAP_WORKFLOW_NOTE),
        ("Risk Category List", "Categories", ", ".join(RISK_CATEGORIES)),
        ("Validation Status Values", "Values", ", ".join(VALIDATION_STATUS_VALUES)),
        ("Engineer Decision Values", "Values", ", ".join(ENGINEER_DECISION_VALUES)),
        ("Action Status Values", "Values", ", ".join(ACTION_STATUS_VALUES)),
        ("Build Phase Values", "Values", ", ".join(BUILD_PHASE_VALUES)),
    ]
    return pd.DataFrame(rows, columns=["Category", "Setting", "Definition"])


def pilot_metrics_dataframe(
    dfmea_df: pd.DataFrame,
    dvp_df: pd.DataFrame,
    trace_df: pd.DataFrame,
    gap_df: pd.DataFrame,
    lessons_df: pd.DataFrame,
) -> pd.DataFrame:
    suggestions = len(dfmea_df) + len(dvp_df) + len(lessons_df)
    accepted = (
        safe_count(dfmea_df, "Engineer Decision", "Accept")
        + safe_count(dvp_df, "Engineer Decision", "Accept")
        + safe_count(lessons_df, "Engineer Decision", "Accept")
    )
    modified = (
        safe_count(dfmea_df, "Engineer Decision", "Modify")
        + safe_count(dvp_df, "Engineer Decision", "Modify")
        + safe_count(lessons_df, "Engineer Decision", "Modify")
    )
    rejected = (
        safe_count(dfmea_df, "Engineer Decision", "Reject")
        + safe_count(dvp_df, "Engineer Decision", "Reject")
        + safe_count(lessons_df, "Engineer Decision", "Reject")
    )
    manual_time = 16.0
    assisted_time = 3.0
    time_saved = round((manual_time - assisted_time) / manual_time, 4) if manual_time else "TBD"
    acceptance_rate: float | str = round(accepted / suggestions, 4) if accepted else "Pending engineer review"
    rows = [
        ("Manual DFMEA/DVP&R Time", "Baseline time entered by engineer", manual_time),
        ("AI-Assisted Time", "Time using tool", assisted_time),
        ("Time Saved %", "Calculated efficiency improvement", time_saved),
        ("AI Suggestions Generated", "Total AI-generated rows", suggestions),
        ("AI Suggestions Accepted", "Engineer accepted suggestions", accepted),
        ("AI Suggestions Modified", "Engineer modified suggestions", modified),
        ("AI Suggestions Rejected", "Engineer rejected suggestions", rejected),
        ("AI Acceptance Rate", "Accepted / generated", acceptance_rate),
        ("Missing Risks Found", "Risks AI found that were not initially listed", len(dfmea_df)),
        ("False Positives", "Suggestions rejected as irrelevant", rejected),
        ("Validation Gaps Found", "Total gaps identified", len(gap_df)),
        ("High-Severity Gaps Found", "Gaps with Severity >= 8", len(gap_df[gap_df["Severity"] >= 8]) if not gap_df.empty else 0),
        ("Lessons Reused", "Lessons learned applied to current component", len(lessons_df)),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Description", "Value"])


def pitch_readiness_checklist_dataframe() -> pd.DataFrame:
    return pd.DataFrame(
        [("[ ]", item, "Review before management pitch") for item in PITCH_READINESS_ITEMS],
        columns=["Complete", "Checklist Item", "Notes"],
    )


def workbook_tables(
    inputs: dict[str, str],
    dfmea_df: pd.DataFrame,
    dvp_df: pd.DataFrame,
    trace_df: pd.DataFrame,
    gap_df: pd.DataFrame,
    lessons_df: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    sorted_dvp_df = sort_dvp_by_test_id(dvp_df)
    return {
        "Management Summary": management_summary_dataframe(inputs, dfmea_df, dvp_df, trace_df, gap_df, lessons_df),
        "Dashboard": dashboard_dataframe(dfmea_df, sorted_dvp_df, trace_df, gap_df, lessons_df),
        "FMEA Header": fmea_header_dataframe(inputs),
        "Component Input": component_input_dataframe(inputs),
        "Knowledge Base Summary": kb_summary_dataframe(),
        "Retrieved Sources": st.session_state.get("retrieved_sources_df", empty_df(RETRIEVED_SOURCES_COLUMNS)),
        "P-Diagram": generate_p_diagram(inputs),
        "DFMEA": dfmea_df,
        "DVP&R": sorted_dvp_df,
        "Traceability": trace_df,
        "Gap Analysis": gap_df,
        "Lessons Learned": lessons_df,
        "Pilot Metrics": pilot_metrics_dataframe(dfmea_df, dvp_df, trace_df, gap_df, lessons_df),
        "Settings": settings_dataframe(),
        "Pitch Readiness Checklist": pitch_readiness_checklist_dataframe(),
    }


def build_report(
    inputs: dict[str, str],
    dfmea_df: pd.DataFrame,
    dvp_df: pd.DataFrame,
    trace_df: pd.DataFrame,
    gap_df: pd.DataFrame,
    lessons_df: pd.DataFrame,
) -> str:
    dashboard = dashboard_dataframe(dfmea_df, dvp_df, trace_df, gap_df, lessons_df)
    percent_metrics = {
        "Estimated RPN Reduction %",
        "Overall Validation Coverage Score",
        "High-Severity Validation Coverage Score",
    }

    def format_kpi(metric: str, value: Any) -> str:
        if metric in percent_metrics and isinstance(value, (int, float)):
            return f"{value:.1%}"
        return str(value)

    kpi_lines = [
        f"- {row['Metric']}: {format_kpi(str(row['Metric']), row['Value'])}"
        for _, row in dashboard[dashboard["Section"] == "KPI"].iterrows()
    ]
    gap_lines = []
    if not gap_df.empty:
        for _, row in gap_df.iterrows():
            gap_lines.append(f"- {row['Gap ID']} {row['Failure Mode']}: {row['Recommended Fix']}")
    else:
        gap_lines.append("- No open validation gaps identified by the current rule set.")

    return f"""# BIW DFMEA-DVP&R Pilot Report

## Engineering Review Disclaimer

{DISCLAIMER}

{PROTOTYPE_BOUNDARY_STATEMENT}

{ACTION_PRIORITY_DISCLAIMER}

## Component Summary

- Component name: {inputs.get("component_name") or "Not specified"}
- Vehicle area: {inputs.get("vehicle_area") or "Not specified"}
- Component category: {inputs.get("component_category") or "Not specified"}
- Material: {inputs.get("material") or "Not specified"}
- Thickness: {inputs.get("thickness") or "Not specified"}
- Joining method: {inputs.get("joining_method") or "Not specified"}
- Manufacturing process: {inputs.get("manufacturing_process") or "Not specified"}
- Program phase: {inputs.get("program_phase") or "Not specified"}
- Engineer name: {inputs.get("engineer_name") or "Not specified"}

## Dashboard KPIs

{chr(10).join(kpi_lines)}

## Demo Story

{DEMO_STORY}

## Dashboard Footnotes

{chr(10).join(f"- {note}" for note in DASHBOARD_FOOTNOTES)}

## Open Gap Items

{chr(10).join(gap_lines)}

{GAP_WORKFLOW_NOTE}

## Recommended Next Steps

- Review all High Action Priority rows with responsible BIW, CAE, validation, manufacturing, quality, and materials owners.
- Confirm DVP&R acceptance criteria and evidence links before closure.
- Update engineer review decisions and approval status after design review.
- Connect future AI/RAG suggestions to approved internal source citations before production use.
"""


def add_dropdown(ws, header_name: str, values: list[str]) -> None:
    headers = [cell.value for cell in ws[1]]
    if header_name not in headers or ws.max_row < 2:
        return
    col = headers.index(header_name) + 1
    col_letter = get_column_letter(col)
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"{col_letter}2:{col_letter}{max(ws.max_row, 500)}")


def apply_status_fills(ws, header_name: str, fill_map: dict[str, PatternFill]) -> None:
    headers = [cell.value for cell in ws[1]]
    if header_name not in headers:
        return
    col = headers.index(header_name) + 1
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=col).value
        fill = fill_map.get(str(value))
        if fill:
            for cell_col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=cell_col).fill = fill


def apply_rpn_range_fills(ws, header_names: list[str], high_fill: PatternFill, medium_fill: PatternFill, low_fill: PatternFill) -> None:
    headers = [cell.value for cell in ws[1]]
    for header_name in header_names:
        if header_name not in headers:
            continue
        col = headers.index(header_name) + 1
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            try:
                rpn = float(value)
            except (TypeError, ValueError):
                continue
            if rpn >= 150:
                ws.cell(row=row, column=col).fill = high_fill
            elif rpn >= 100:
                ws.cell(row=row, column=col).fill = medium_fill
            else:
                ws.cell(row=row, column=col).fill = low_fill


def set_series_titles(chart: BarChart | PieChart, titles: list[str]) -> None:
    for series, title in zip(chart.series, titles):
        series.tx = SeriesLabel(v=title)


def add_dashboard_charts(workbook) -> None:
    ws = workbook["Dashboard"]
    sections: dict[str, tuple[int, int]] = {}
    current_section = ""
    current_start = 0
    for row in range(2, ws.max_row + 1):
        section = str(ws.cell(row=row, column=1).value or "")
        if section.endswith("Header"):
            if current_section:
                sections[current_section] = (current_start, row - 1)
            current_section = section
            current_start = row
    if current_section:
        sections[current_section] = (current_start, ws.max_row)

    if "RPN Chart Header" in sections:
        start, end = sections["RPN Chart Header"]
        chart = BarChart()
        chart.title = "Initial RPN vs Revised RPN"
        chart.y_axis.title = "RPN"
        chart.x_axis.title = "Failure Mode"
        data = Reference(ws, min_col=3, max_col=4, min_row=start, max_row=end)
        cats = Reference(ws, min_col=2, min_row=start + 1, max_row=end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        set_series_titles(chart, ["Initial RPN", "Revised RPN"])
        chart.height = 7
        chart.width = 16
        ws.add_chart(chart, "F2")

    if "RPN Chart Header" in sections:
        start, end = sections["RPN Chart Header"]
        chart = BarChart()
        chart.title = "RPN by Failure Mode"
        chart.y_axis.title = "Initial RPN"
        chart.x_axis.title = "Failure Mode"
        data = Reference(ws, min_col=3, min_row=start, max_row=end)
        cats = Reference(ws, min_col=2, min_row=start + 1, max_row=end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        set_series_titles(chart, ["Initial RPN"])
        chart.height = 7
        chart.width = 15
        ws.add_chart(chart, "F18")

    if "Coverage Score Header" in sections:
        start, end = sections["Coverage Score Header"]
        chart = BarChart()
        chart.title = "Validation Coverage Score"
        chart.y_axis.title = "Coverage Score"
        chart.x_axis.title = "Coverage Measure"
        data = Reference(ws, min_col=3, min_row=start, max_row=end)
        cats = Reference(ws, min_col=2, min_row=start + 1, max_row=end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        set_series_titles(chart, ["Validation Coverage Score"])
        chart.height = 7
        chart.width = 15
        ws.add_chart(chart, "F34")

    if "Coverage Status Header" in sections:
        start, end = sections["Coverage Status Header"]
        chart = PieChart()
        chart.title = "Validation Coverage Status"
        data = Reference(ws, min_col=3, min_row=start, max_row=end)
        labels = Reference(ws, min_col=2, min_row=start + 1, max_row=end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        set_series_titles(chart, ["Coverage Status"])
        chart.height = 7
        chart.width = 10
        ws.add_chart(chart, "F50")

    if "Risk Category Header" in sections:
        start, end = sections["Risk Category Header"]
        chart = BarChart()
        chart.title = "Failure Modes by Risk Category"
        chart.y_axis.title = "Failure Mode Count"
        chart.x_axis.title = "Risk Category"
        data = Reference(ws, min_col=3, min_row=start, max_row=end)
        cats = Reference(ws, min_col=2, min_row=start + 1, max_row=end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        set_series_titles(chart, ["Failure Mode Count"])
        chart.height = 7
        chart.width = 15
        ws.add_chart(chart, "P2")

    if "Open Actions Header" in sections:
        start, end = sections["Open Actions Header"]
        chart = BarChart()
        chart.title = "Open Actions by Owner"
        chart.y_axis.title = "Open Action Count"
        chart.x_axis.title = "Owner"
        data = Reference(ws, min_col=3, min_row=start, max_row=end)
        cats = Reference(ws, min_col=2, min_row=start + 1, max_row=end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        set_series_titles(chart, ["Open Action Count"])
        chart.height = 7
        chart.width = 15
        ws.add_chart(chart, "P18")


def format_excel_workbook(writer: pd.ExcelWriter, tables: dict[str, pd.DataFrame]) -> None:
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill("solid", fgColor="F4CCCC")
    medium_fill = PatternFill("solid", fgColor="FFF2CC")
    low_fill = PatternFill("solid", fgColor="D9EAD3")
    partial_fill = PatternFill("solid", fgColor="FFF2CC")
    proposed_fill = PatternFill("solid", fgColor="FCE4D6")
    accepted_fill = PatternFill("solid", fgColor="D9EAF7")
    waived_fill = PatternFill("solid", fgColor="D9D9D9")
    failed_fill = PatternFill("solid", fgColor="F4CCCC")
    draft_fill = PatternFill("solid", fgColor="EADCF8")
    card_fill = PatternFill("solid", fgColor="EAF2F8")
    prototype_fill = PatternFill("solid", fgColor="FCE4D6")

    for sheet_name, df in tables.items():
        ws = workbook[sheet_name[:31]]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        apply_status_fills(ws, "Action Priority", {"High": high_fill, "Medium": medium_fill, "Low": low_fill})
        apply_status_fills(ws, "AP (AIAG-VDA)", {"H": high_fill, "M": medium_fill, "L": low_fill})
        apply_status_fills(
            ws,
            "Special Characteristic",
            {"YC (Potential Critical)": high_fill, "SC (Potential Significant)": medium_fill},
        )
        apply_status_fills(ws, "Residual Risk Level", {"High": high_fill, "Medium": medium_fill, "Low": low_fill})
        apply_status_fills(
            ws,
            "Coverage Status",
            {"Gap": high_fill, "Partial": partial_fill, "Proposed Coverage": proposed_fill, "Covered": low_fill},
        )
        apply_status_fills(
            ws,
            "Gap Status",
            {
                "Open": high_fill,
                "Proposed": proposed_fill,
                "Accepted": accepted_fill,
                "In Progress": medium_fill,
                "Closed": low_fill,
                "Waived": waived_fill,
            },
        )
        apply_status_fills(
            ws,
            "Gap Closure Status",
            {
                "Open": high_fill,
                "Proposed": proposed_fill,
                "Accepted": accepted_fill,
                "In Progress": medium_fill,
                "Closed": low_fill,
                "Waived": waived_fill,
            },
        )
        apply_status_fills(
            ws,
            "Final Gap Closure Status",
            {
                "Open": high_fill,
                "Proposed": proposed_fill,
                "Accepted": accepted_fill,
                "In Progress": medium_fill,
                "Closed": low_fill,
                "Waived": waived_fill,
            },
        )
        apply_status_fills(
            ws,
            "Validation Status",
            {
                "Planned": accepted_fill,
                "Proposed": proposed_fill,
                "In Progress": medium_fill,
                "Passed": low_fill,
                "Failed": failed_fill,
                "Waived": waived_fill,
                "Blocked": high_fill,
            },
        )
        apply_status_fills(ws, "Approval Status", {"Draft": draft_fill})
        apply_rpn_range_fills(ws, ["RPN", "Initial RPN", "Revised RPN"], high_fill, medium_fill, low_fill)

        add_dropdown(ws, "Action Status", ACTION_STATUS_VALUES)
        add_dropdown(ws, "Validation Level", VALIDATION_LEVEL_VALUES)
        add_dropdown(ws, "Build Phase", BUILD_PHASE_VALUES)
        add_dropdown(ws, "Test Stage", TEST_STAGE_VALUES)
        add_dropdown(ws, "Special Characteristic", [v for v in SPECIAL_CHARACTERISTIC_VALUES if v])
        add_dropdown(ws, "Pass / Fail", PASS_FAIL_VALUES)
        add_dropdown(ws, "Validation Status", VALIDATION_STATUS_VALUES)
        add_dropdown(ws, "Gap Status", GAP_STATUS_VALUES)
        add_dropdown(ws, "Gap Closure Status", GAP_STATUS_VALUES)
        add_dropdown(ws, "Final Gap Closure Status", GAP_STATUS_VALUES)
        add_dropdown(ws, "Engineer Decision", ENGINEER_DECISION_VALUES)
        add_dropdown(ws, "Approval Status", APPROVAL_STATUS_VALUES)
        add_dropdown(ws, "Rejection Reason", REJECTION_REASON_VALUES)

        headers = [cell.value for cell in ws[1]]
        percent_headers = {"RPN Reduction %"}
        for percent_header in percent_headers:
            if percent_header in headers and ws.max_row >= 2:
                col = headers.index(percent_header) + 1
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).number_format = "0.0%"

        for numeric_header in ["RPN", "Initial RPN", "Revised RPN"]:
            if numeric_header in headers and ws.max_row >= 2:
                col = headers.index(numeric_header) + 1
                letter = get_column_letter(col)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    ColorScaleRule(start_type="min", start_color="D9EAD3", mid_type="percentile", mid_value=50, mid_color="FFF2CC", end_type="max", end_color="F4CCCC"),
                )

        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 60))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)

        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 42 if row_idx > 1 else 30

        if sheet_name == "Dashboard":
            percent_metrics = {
                "Estimated RPN Reduction %",
                "Overall Validation Coverage Score",
                "High-Severity Validation Coverage Score",
            }
            for row in range(2, ws.max_row + 1):
                section = ws.cell(row=row, column=1).value
                metric = ws.cell(row=row, column=2).value
                if section in ("KPI", "RAG KPI"):
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = card_fill
                    ws.cell(row=row, column=2).font = Font(bold=True)
                    ws.cell(row=row, column=3).font = Font(bold=True, size=14)
                if section == "Prototype Boundary":
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = prototype_fill
                if section == "Dashboard Footnote":
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = low_fill
                if section == "Coverage Score":
                    ws.cell(row=row, column=3).number_format = "0.0%"
                if metric in percent_metrics:
                    ws.cell(row=row, column=3).number_format = "0.0%"
            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 42
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 18

        if sheet_name == "Pilot Metrics":
            for row in range(2, ws.max_row + 1):
                metric = ws.cell(row=row, column=1).value
                if metric in {"Time Saved %", "AI Acceptance Rate"} and isinstance(ws.cell(row=row, column=3).value, (int, float)):
                    ws.cell(row=row, column=3).number_format = "0.0%"

        if sheet_name == "Management Summary":
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 110
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=1).fill = card_fill
                ws.cell(row=row, column=1).font = Font(bold=True)

        if sheet_name == "Pitch Readiness Checklist":
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 48
            ws.column_dimensions["C"].width = 36
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=1).fill = card_fill

    add_dashboard_charts(workbook)


def workbook_bytes(tables: dict[str, pd.DataFrame]) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        format_excel_workbook(writer, tables)
    buffer.seek(0)
    return buffer.getvalue()


def load_part_example(label: str) -> None:
    example = PART_EXAMPLES.get(label, {})
    for key, value in example.items():
        st.session_state[key] = value
    st.session_state.loaded_part_label = label


def selected_part_label() -> str:
    label = str(st.session_state.get("selected_part_label", "")).strip()
    if label in PART_EXAMPLES:
        return label
    return DEFAULT_PART_LABEL


APP_CSS = """
<style>
/* ---- global typography and spacing ---- */
.block-container { padding-top: 1.2rem; }

/* ---- hero banner ---- */
.biw-hero {
    background: linear-gradient(135deg, #10312B 0%, #1A4A40 70%, #17423A 100%);
    border-radius: 14px;
    padding: 1.5rem 1.8rem 1.35rem 1.8rem;
    margin-bottom: 1.0rem;
}
.biw-hero h1 {
    color: #FFFFFF;
    font-size: 1.65rem;
    font-weight: 700;
    margin: 0 0 0.25rem 0;
    padding: 0;
}
.biw-hero p {
    color: #C9DAD4;
    font-size: 0.92rem;
    margin: 0;
}
.biw-badge {
    display: inline-block;
    background: #E8A33D;
    color: #10312B;
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.08em;
    border-radius: 999px;
    padding: 0.15rem 0.7rem;
    margin-bottom: 0.55rem;
    text-transform: uppercase;
}
.biw-disclaimer {
    background: rgba(232, 163, 61, 0.14);
    border-radius: 8px;
    color: #F4D9A8;
    font-size: 0.8rem;
    padding: 0.45rem 0.75rem;
    margin-top: 0.75rem;
}

/* ---- tabs ---- */
.stTabs [data-baseweb="tab-list"] { gap: 0.35rem; }
.stTabs [data-baseweb="tab"] {
    background: #EEF4F1;
    border-radius: 8px 8px 0 0;
    padding: 0.5rem 1.0rem;
    font-weight: 600;
}
.stTabs [aria-selected="true"] {
    background: #10312B !important;
    color: #FFFFFF !important;
}

/* ---- metric cards ---- */
div[data-testid="stMetric"] {
    background: #EEF4F1;
    border-radius: 10px;
    padding: 0.7rem 0.9rem;
    box-shadow: 0 1px 4px rgba(16, 49, 43, 0.10);
}
div[data-testid="stMetric"] label { color: #5C6B66; }

/* ---- buttons ---- */
.stButton > button[kind="primary"], .stDownloadButton > button {
    border-radius: 8px;
    font-weight: 600;
}
</style>
"""


def render_header() -> None:
    st.markdown(APP_CSS, unsafe_allow_html=True)
    st.markdown(
        f"""
        <div class="biw-hero">
            <span class="biw-badge">Prototype {TOOL_VERSION} · AIAG-VDA aligned · RAG grounded · Engineer approved</span>
            <h1>{APP_TITLE}</h1>
            <p>Component inputs &rarr; knowledge retrieval &rarr; P-Diagram &rarr; draft DFMEA &rarr; recommended DVP&amp;R &rarr; traceability &amp; gap detection &rarr; cited, management-ready export.</p>
            <div class="biw-disclaimer">&#9888;&#65039; {DISCLAIMER}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.subheader("Prototype Mode")
        st.caption("Rules-based generator is active. OpenAI and RAG can be added later as generation providers.")
        if PART_EXAMPLES:
            labels = list(PART_EXAMPLES)
            current_label = selected_part_label()
            index = labels.index(current_label) if current_label in labels else 0
            chosen = st.selectbox(
                "Example part",
                labels,
                index=index,
                key="selected_part_label",
                help="Example parts are loaded from examples/parts/*.json.",
            )
            if st.button("Load Selected Part", width="stretch"):
                load_part_example(chosen)
                st.toast(f"Loaded {chosen}.")
        else:
            st.error("No example part files found in examples/parts.")

        if RAG_IMPORT_ERROR:
            st.caption("RAG layer unavailable - rules-based generation only.")
        else:
            llm_ok, llm_state = llm_status()
            st.checkbox(
                "Enable LLM enrichment (optional)",
                key="llm_enrich_mode",
                disabled=not llm_ok,
                help=(
                    "Off by default. When a provider is configured via .env (RAG_LLM_PROVIDER=anthropic or "
                    "openai), the app sends retrieved context to the LLM for additional schema-validated, "
                    "source-cited suggestions. All LLM rows require engineer review."
                ),
            )
            st.caption(f"LLM provider: {llm_state}")
            for note in st.session_state.get("llm_notes", [])[:3]:
                st.caption(f"LLM note: {note}")
        st.checkbox(
            "Show intentional demo gap",
            key="demo_gap_mode",
            help="When enabled, a bracket or fastener risk may be left without validation so gap detection is visible. Crash validation is always generated.",
        )
        st.divider()
        if st.button("Generate Drafts", type="primary", width="stretch"):
            generate_all()
            st.toast("Drafts generated.")


def render_input() -> None:
    st.subheader("Input")
    if PART_EXAMPLES:
        labels = list(PART_EXAMPLES)
        current_label = selected_part_label()
        index = labels.index(current_label) if current_label in labels else 0
        col_part, col_load = st.columns([3, 1])
        with col_part:
            chosen = st.selectbox(
                "Load example part",
                labels,
                index=index,
                key="component_tab_part_label",
                help="Add more JSON files to examples/parts to expand this list.",
            )
        with col_load:
            st.write("")
            st.write("")
            if st.button("Load Part"):
                load_part_example(chosen)
                st.success(f"Loaded {chosen}.")

    left, right = st.columns(2)
    with left:
        st.text_input("Component name", key="component_name")
        st.text_input("Vehicle area", key="vehicle_area")
        st.text_input("Component category", key="component_category")
        st.text_input("Material", key="material")
        st.text_input("Thickness", key="thickness")
        st.text_input("Joining method", key="joining_method")
        st.text_input("Program phase", key="program_phase")
        st.text_input("Engineer name", key="engineer_name")

    with right:
        st.text_input("Manufacturing process", key="manufacturing_process")
        st.text_area("Function", key="primary_function", height=80)
        st.text_area("Interfaces", key="interfaces", height=80)
        st.text_area("Load cases", key="load_cases", height=80)
        st.text_area("Environmental exposure", key="environmental_exposure", height=80)
        st.text_area("Known concerns", key="known_design_concerns", height=80)

    st.text_area("Applicable assumptions", key="assumptions", height=70)

    with st.expander("FMEA header (AIAG-VDA planning & preparation)", expanded=False):
        st.caption("Auto-derived identification block. A production version would pull this from the program FMEA register.")
        st.dataframe(fmea_header_dataframe(component_inputs()), width="stretch", hide_index=True)

    if st.button("Generate Drafts", type="primary"):
        generate_all()
        st.success("P-Diagram, draft DFMEA, DVP&R, traceability, gaps, dashboard, and lessons generated.")


def render_knowledge_base() -> None:
    st.subheader("Knowledge Base (Local RAG)")
    st.caption(RAG_DISCLAIMER)
    if RAG_IMPORT_ERROR:
        st.error(f"RAG layer unavailable: {RAG_IMPORT_ERROR}. Install requirements and restart.")
        return
    store = get_rag_store()
    stats = store.get_collection_stats()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Documents Indexed", stats["documents"])
    c2.metric("Knowledge Chunks", stats["chunks"])
    c3.metric("Document Types", len(stats["document_types"]))
    c4.metric("Embedding", "Semantic" if "MiniLM" in stats["embedding_model"] else "Fallback")
    st.caption(f"Embedding model: {stats['embedding_model']} · Store: {stats['path']}")

    col_seed, col_clear = st.columns(2)
    with col_seed:
        if st.button("Seed synthetic demo corpus", help="Index the bundled prior-program DFMEA/DVP&R/lessons/standards demo files."):
            added = seed_knowledge_base(store)
            st.success(f"Indexed {added} new chunks from the synthetic corpus.")
            st.rerun()
    with col_clear:
        if st.button("Clear knowledge base", help="Delete all indexed chunks and embeddings."):
            store.delete_collection()
            st.success("Knowledge base cleared.")
            st.rerun()

    st.divider()
    st.markdown("**Upload engineering documents** (synthetic or approved non-confidential files only)")
    up1, up2 = st.columns(2)
    with up1:
        doc_type = st.selectbox("Document type", RAG_DOC_TYPES, key="kb_doc_type")
        component_hint = st.text_input(
            "Component type (if known)",
            key="kb_component_hint",
            placeholder="e.g. Front rail reinforcement",
        )
    with up2:
        strength = st.selectbox("Source strength", RAG_SOURCE_STRENGTHS, key="kb_source_strength")
        upload_notes = st.text_input("Optional notes", key="kb_upload_notes", placeholder="e.g. anonymized P1 program data")
    uploads = st.file_uploader(
        "Files (.xlsx, .csv, .md, .txt, .pdf, .docx)",
        type=["xlsx", "csv", "md", "txt", "pdf", "docx"],
        accept_multiple_files=True,
        key="kb_uploads",
    )
    if uploads and st.button("Process files", type="primary"):
        KB_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        total = 0
        for file in uploads:
            saved = KB_UPLOAD_DIR / file.name
            saved.write_bytes(file.getvalue())
            chunks = load_file_to_chunks(saved, doc_type, strength, file_bytes=file.getvalue())
            for chunk in chunks:
                if component_hint and not chunk["metadata"].get("component_name"):
                    chunk["metadata"]["component_name"] = component_hint
                if upload_notes:
                    chunk["metadata"]["notes"] = upload_notes
            total += store.add_chunks(chunks)
        st.success(f"Indexed {total} new chunks from {len(uploads)} file(s). Duplicates were skipped.")
        st.rerun()

    st.divider()
    st.markdown("**Search preview** — check what the assistant would retrieve")
    query = st.text_input("Query", key="kb_query", placeholder="e.g. spot weld fatigue on high-strength steel rail flange")
    if query:
        results = store.search(query, top_k=5)
        if results:
            preview = pd.DataFrame(
                [
                    {
                        "Similarity": r["similarity"],
                        "Document Type": r["metadata"].get("document_type", ""),
                        "Source File": r["metadata"].get("file_name", ""),
                        "Sheet": r["metadata"].get("sheet_name", ""),
                        "Row": r["metadata"].get("row_number", ""),
                        "Preview": r["chunk_text"][:160],
                    }
                    for r in results
                ]
            )
            st.dataframe(preview, width="stretch", hide_index=True)
        else:
            st.info("No results — seed the corpus or upload documents first.")

    if not st.session_state.retrieved_sources_df.empty:
        st.divider()
        st.markdown("**Sources retrieved during the last generation**")
        st.dataframe(st.session_state.retrieved_sources_df, width="stretch", hide_index=True)


def render_pdiagram() -> None:
    st.subheader("P-Diagram (Function Analysis)")
    st.caption(
        "Parameter diagram supporting AIAG-VDA Step 3: ideal function, input signal, control factors, "
        "the five noise factor categories, and error states that feed the failure analysis."
    )
    if st.button("Generate P-Diagram"):
        st.session_state.pdiag_df = generate_p_diagram(component_inputs())
        st.success("P-Diagram generated from component inputs.")
    if st.session_state.pdiag_df.empty:
        st.info("Load a part and click Generate Drafts (or Generate P-Diagram) to build the function analysis.")
        return
    pdiag = st.session_state.pdiag_df
    c1, c2, c3 = st.columns(3)
    c1.metric("Control Factors", int((pdiag["Element"] == "Control Factor").sum()))
    c2.metric("Noise Factor Categories", int(pdiag["Element"].str.startswith("Noise:").sum()))
    c3.metric("Error States", int((pdiag["Element"] == "Error State").sum()))
    st.session_state.pdiag_df = st.data_editor(
        pdiag,
        num_rows="dynamic",
        width="stretch",
        hide_index=True,
    )
    st.caption("Error states listed here should each map to a failure mode in the DFMEA tab.")
    st.download_button("Download P-Diagram CSV", dataframe_to_csv(st.session_state.pdiag_df), "p_diagram.csv", "text/csv")


def render_dfmea() -> None:
    st.subheader("DFMEA")
    st.caption(
        "Includes stable IDs, vehicle-level end effects, special characteristic candidates (YC/SC), S/O/D, "
        "RPN, AIAG-VDA Action Priority, residual risk, action tracking, and engineer review fields."
    )
    if st.button("Generate DFMEA"):
        st.session_state.dfmea_df = generate_dfmea(component_inputs())
        st.session_state.dvp_df = empty_df(DVP_COLUMNS)
        st.session_state.trace_df = empty_df(TRACE_COLUMNS)
        st.session_state.gap_df = empty_df(GAP_COLUMNS)
        st.session_state.lessons_df = generate_lessons(st.session_state.dfmea_df)
        st.success("DFMEA generated. Generate DVP&R next to create linked validation rows.")
    st.session_state.dfmea_df = st.data_editor(
        st.session_state.dfmea_df,
        num_rows="dynamic",
        width="stretch",
        hide_index=True,
        column_config={
            "Initial Severity": st.column_config.NumberColumn(min_value=1, max_value=10, step=1),
            "Initial Occurrence": st.column_config.NumberColumn(min_value=1, max_value=10, step=1),
            "Initial Detection": st.column_config.NumberColumn(min_value=1, max_value=10, step=1),
            "Severity": st.column_config.NumberColumn(min_value=1, max_value=10, step=1),
            "Occurrence": st.column_config.NumberColumn(min_value=1, max_value=10, step=1),
            "Detection": st.column_config.NumberColumn(min_value=1, max_value=10, step=1),
            "RPN Reduction %": st.column_config.NumberColumn(format="percent"),
            "Special Characteristic": st.column_config.SelectboxColumn(options=SPECIAL_CHARACTERISTIC_VALUES),
            "AP (AIAG-VDA)": st.column_config.SelectboxColumn(options=["H", "M", "L"]),
        },
    )
    refresh_downstream_from_edits()
    st.download_button("Download DFMEA CSV", dataframe_to_csv(st.session_state.dfmea_df), "dfmea_draft.csv", "text/csv")


def render_dvp() -> None:
    st.subheader("DVP&R")
    st.caption("Expanded validation plan with test IDs, validation level, build phase, status, results, and evidence tracking.")
    if st.button("Generate DVP&R"):
        if st.session_state.dfmea_df.empty:
            st.session_state.dfmea_df = generate_dfmea(component_inputs())
        st.session_state.dvp_df = generate_dvp(st.session_state.dfmea_df, generation_settings())
        st.session_state.trace_df = generate_traceability(st.session_state.dfmea_df, st.session_state.dvp_df)
        st.session_state.gap_df = generate_gap_analysis(st.session_state.trace_df)
        st.success("DVP&R generated and linked to DFMEA failure mode IDs.")
    st.session_state.dvp_df = st.data_editor(
        st.session_state.dvp_df,
        num_rows="dynamic",
        width="stretch",
        hide_index=True,
    )
    refresh_downstream_from_edits()
    st.download_button("Download DVP&R CSV", dataframe_to_csv(st.session_state.dvp_df), "dvpr_draft.csv", "text/csv")


def render_traceability() -> None:
    st.subheader("Traceability")
    if st.button("Check Gaps"):
        refresh_downstream_from_edits()
        st.success("Traceability and gap analysis refreshed.")
    else:
        refresh_downstream_from_edits()
    trace_df = st.session_state.trace_df
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Failure Modes", len(trace_df))
    c2.metric("Covered", safe_count(trace_df, "Coverage Status", "Covered"))
    c3.metric("Partial", safe_count(trace_df, "Coverage Status", "Partial"))
    c4.metric("Gaps", safe_count(trace_df, "Coverage Status", "Gap"))
    c5.metric("Coverage Score", f"{mean_numeric(trace_df, 'Coverage Score'):.0f}%")
    st.dataframe(trace_df, width="stretch", hide_index=True)
    st.download_button("Download Traceability CSV", dataframe_to_csv(trace_df), "traceability.csv", "text/csv")


def render_gaps() -> None:
    st.subheader("Gaps")
    col_a, col_b = st.columns(2)
    with col_a:
        if st.button("Refresh Gap Analysis"):
            refresh_downstream_from_edits()
            st.success("Gap analysis refreshed.")
    with col_b:
        if st.button("Generate Lessons Learned"):
            st.session_state.lessons_df = generate_lessons(st.session_state.dfmea_df)
            st.success("Lessons learned regenerated.")
    refresh_downstream_from_edits()
    st.dataframe(st.session_state.gap_df, width="stretch", hide_index=True)
    st.download_button("Download Gap Analysis CSV", dataframe_to_csv(st.session_state.gap_df), "gap_analysis.csv", "text/csv")
    st.subheader("Lessons Learned")
    st.dataframe(st.session_state.lessons_df, width="stretch", hide_index=True)
    st.download_button("Download Lessons CSV", dataframe_to_csv(st.session_state.lessons_df), "lessons.csv", "text/csv")


def render_dashboard() -> None:
    refresh_downstream_from_edits()
    st.subheader("Dashboard")
    dashboard = dashboard_dataframe(
        st.session_state.dfmea_df,
        st.session_state.dvp_df,
        st.session_state.trace_df,
        st.session_state.gap_df,
        st.session_state.lessons_df,
    )
    kpis = dashboard[dashboard["Section"] == "KPI"]
    metric_cols = st.columns(4)
    for idx, (_, row) in enumerate(kpis.head(8).iterrows()):
        metric_cols[idx % 4].metric(str(row["Metric"]), str(row["Value"]))

    # RAG grounding KPIs
    if not RAG_IMPORT_ERROR:
        try:
            stats = get_rag_store().get_collection_stats()
            dfmea_df, dvp_df = st.session_state.dfmea_df, st.session_state.dvp_df
            grounded = safe_count(dfmea_df, "Source Type", "RAG Retrieved (rules draft)") + safe_count(
                dvp_df, "Source Type", "RAG Retrieved (rules draft)"
            )
            total_rows = len(dfmea_df) + len(dvp_df)
            fallback = total_rows - grounded
            avg_conf = 0.0
            if total_rows:
                avg_conf = (
                    mean_numeric(dfmea_df, "AI Confidence Score") * len(dfmea_df)
                    + mean_numeric(dvp_df, "AI Confidence Score") * len(dvp_df)
                ) / total_rows
            st.markdown("**RAG knowledge layer**")
            r1, r2, r3, r4, r5 = st.columns(5)
            r1.metric("Documents Indexed", stats["documents"])
            r2.metric("Knowledge Chunks", stats["chunks"])
            r3.metric("Rows with Source Evidence", grounded)
            r4.metric("Rule-Based Fallback Rows", fallback)
            r5.metric("Avg AI Confidence", f"{avg_conf:.2f}")
        except Exception:
            pass
    st.dataframe(dashboard.astype(str), width="stretch", hide_index=True)

    if not st.session_state.dfmea_df.empty:
        st.bar_chart(st.session_state.dfmea_df["Risk Category"].value_counts())
    if not st.session_state.trace_df.empty:
        st.bar_chart(st.session_state.trace_df["Coverage Status"].value_counts())


def render_export() -> None:
    refresh_downstream_from_edits()
    st.subheader("Export")
    inputs = component_inputs()
    tables = workbook_tables(
        inputs,
        st.session_state.dfmea_df,
        st.session_state.dvp_df,
        st.session_state.trace_df,
        st.session_state.gap_df,
        st.session_state.lessons_df,
    )
    report = build_report(
        inputs,
        st.session_state.dfmea_df,
        st.session_state.dvp_df,
        st.session_state.trace_df,
        st.session_state.gap_df,
        st.session_state.lessons_df,
    )
    st.markdown(report)
    excel_data = workbook_bytes(tables)
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "Download Excel Workbook",
            excel_data,
            "biw_dfmea_dvpr_outputs.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col2:
        st.download_button(
            "Download Markdown Report",
            report.encode("utf-8"),
            "biw_dfmea_dvpr_pilot_report.md",
            "text/markdown",
        )


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="🚙", layout="wide")
    init_state()
    render_header()

    tabs = st.tabs(["Input", "Knowledge Base", "P-Diagram", "DFMEA", "DVP&R", "Traceability", "Gaps", "Dashboard", "Export"])
    with tabs[0]:
        render_input()
    with tabs[1]:
        render_knowledge_base()
    with tabs[2]:
        render_pdiagram()
    with tabs[3]:
        render_dfmea()
    with tabs[4]:
        render_dvp()
    with tabs[5]:
        render_traceability()
    with tabs[6]:
        render_gaps()
    with tabs[7]:
        render_dashboard()
    with tabs[8]:
        render_export()


if __name__ == "__main__":
    main()
